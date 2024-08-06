from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from datetime import datetime
from bs4 import BeautifulSoup
import pandas as pd
import time
import os
import configparser
import threading
from openpyxl import load_workbook
import openpyxl
from base.mystruct import TreeNode
from base.boot import *
from base.waitclick import *
from base.membase import *


stop_event = threading.Event()
amount_that_complete = 0
org_directory = ""
captcha = ""


def stop_org_thread():
    stop_event.set()



def load_config(self):
        if os.path.exists(self.config_file):
            self.config.read(self.config_file)
            self.explore_driver_path.insert(0, self.config.get('Paths', 'explore_driver_path', fallback=''))



def access_org_database(driver, wait):
    while(1):
        try:
            Databaseofparty = wait.until(EC.element_to_be_clickable((By.XPATH, '(//img[contains(@src, "党组织和党员信息库.png")])[2]')))
            break
        except:
            time.sleep(0.5)
    while(1):
        try:
            Databaseofparty.click()
            break
        except:

            Databaseofparty = wait.until(EC.element_to_be_clickable((By.XPATH, '(//img[contains(@src, "党组织和党员信息库.png")])[2]')))
            time.sleep(0.5)

    try:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if '党组织和党员信息' in driver.title:
                break
        print(f"进入党组织和党员信息库成功")
    except:
        time.sleep(1)



def switch_role(wait):
    while 1:
        try:
            droplist = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.avatar-wrapper.fs-dropdown-selfdefine')))
            droplist.click()
            print(f"进入角色下拉列表成功")
            time.sleep(1)
            config_file = 'config.ini'
            config = configparser.ConfigParser()
            config.read(config_file)
            role_name_org = config.get('role_org_name', 'name_role_org', fallback='')
            #print(f'//SPAN[contains(text(), "{role_name_org}")]')
            role = wait.until(EC.element_to_be_clickable((By.XPATH, f'//SPAN[contains(text(), "{role_name_org}")]')))
            role.click()
            if wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.avatar-wrapper.fs-dropdown-selfdefine'))).text == f"{role_name_org}":
                print(f"切换角色成功")
                break
            else:
                pass
        except:
            time.sleep(2)



#切换到党组织信息页面
def switch_item_org(wait):
    while(1):
        try:
            item_org = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "党组织管理")]/..')))
            break
        except:
            time.sleep(0.1)
    while(1):
        try:
            item_org.click()
            break
        except:
            item_org = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[contains(text(), "党组织管理")]/..')))
    while(1):
        try:
            org_info = wait.until(EC.element_to_be_clickable((By.XPATH, '(//*[contains(text(), "信息管理")]/..)[3]')))
            break
        except:
            time.sleep(0.1)
    while(1):
        try:
            org_info.click()
            break
        except:
            org_info = wait.until(EC.element_to_be_clickable((By.XPATH, '(//*[contains(text(), "信息管理")]/..)[3]')))


def rebuild(driver, wait, excel_file_path, org_excel, ws, wb, excel_file_path_m):
    global amount_that_complete
    #！！！！！@
    driver.refresh()
    time.sleep(3)
    sheet = ws
    amount_that_complete = init_complete_amount(excel_file_path)
    org_totol_amount = count_non_empty_rows(excel_file_path_m, sheet_name=0)

    config = configparser.ConfigParser()
    config.read('config.ini')
    org_directory = config.get('Paths_org_info', 'org_info_path')
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党组织信息库.xlsx"
    excel_file_path_m = os.path.join(org_directory, excel_file_name)
    workbook = openpyxl.load_workbook(excel_file_path_m)
    sheet0 = workbook.active



    for j in range(amount_that_complete, org_totol_amount+1):
        if amount_that_complete%20 == 0:
            driver.close()
            try:
                for handle in driver.window_handles:
                    driver.switch_to.window(handle)
                    if '管理员工作台' in driver.title:
                        break
            except:
                time.sleep(1)
            access_org_database(driver, wait)
            switch_role(wait)
            switch_item_org(wait)
            time.sleep(3)

        string = sheet.cell(row = j, column = 3).value
        array = string.split('/')
        array = array[1:]
        #！！！！！@  从excel中读取
        current_node_text = array[-1]    
        container = wait_return_subelement_absolute(wait, 1, xpath="//div[@class = 'tree_wrapper_div']")
        
        for i, node_value in enumerate(array):
            complete_path = wait_return_subelement_relative(1, container, xpath=f".//span[contains(text(), '{node_value}')]/../../..")
            arrow = wait_return_subelement_relative(1, complete_path, xpath="./span")
            html = complete_path.get_attribute("outerHTML")
            soup_html = BeautifulSoup(html, 'html.parser')
            span_element = soup_html.find('span')
            t = span_element.get('class')
            if span_element and 'expanded' in t:
                continue
            else:
                while 1:
                    try:
                        arrow.click()
                        break
                    except:
                        arrow = wait_return_subelement_relative(1, complete_path, xpath="./span")

        target = wait_return_subelement_relative_v2(1, container, xpath=f".//span[contains(text(), '{current_node_text}')]/..")
        target.click()
        downloading(file = org_excel, wait = wait, driver = driver, path = excel_file_path, rebuild = True, ws=sheet0, wb=workbook)
        

def init(driver, wait, org_directory):
    time.sleep(3)
    scroll = wait_return_subelement_absolute(wait, 1, "//div[@class = 'tree_wrapper']")
    driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scroll)
    time.sleep(2)
    org_tree = TreeNode()
    synchronizing_org_v1(driver=driver, wait=wait, input_node=org_tree, xpath = "(//div[@class = 'fs-tree-node is-expanded is-current is-focusable']/div)[1]", xpath2 = "//div[@role = 'group' and @class = 'fs-tree-node__children']", xpath3 = ".//span[@class = 'fs-tree-node__expand-icon fs-icon-caret-right']", xpath4 = ".//div[@role = 'group']")
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党组织名单.xlsx"
    excel_file_path = os.path.join(org_directory, excel_file_name)
    if os.path.isfile(excel_file_path):
        print("党组织名录已存在!")
        return False
    else:
        os.makedirs(org_directory, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sheet1"
        node_amount = org_tree.return_node_amount()
        for i in range(1, node_amount+1):
            cur_node = org_tree.find_node_by_index(i)
            value = cur_node.value
            ws.cell(row = i, column = 1, value=i)
            ws.cell(row = i, column = 2, value=value)
            path_to_node = cur_node.path_to_root()
            path_to_node = path_to_node[0:]
            char = ''
            for element in path_to_node:
                char = char + '/' + element
            ws.cell(row = i, column = 3, value=char)
        wb.save(excel_file_path)
        return node_amount



def new_excel(wait, driver):
    global org_directory
    config = configparser.ConfigParser()
    config.read('config.ini')
    org_directory = config.get('Paths_org_info', 'org_info_path')
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党组织信息库.xlsx"
    excel_file_path = os.path.join(org_directory, excel_file_name)
    #！！！！！初始化，将党组织信息录入到excel中
    
    if os.path.isfile(excel_file_path):
        time.sleep(3)
        org_excel = load_workbook(excel_file_path)
        config = configparser.ConfigParser()
        config.read('config.ini')
        org_directory = config.get('Paths_org_info', 'org_info_path')
        today_date = datetime.now().strftime("%Y-%m-%d")
        excel_file_name = f"{today_date}党组织名单.xlsx"
        excel_file_path_m = os.path.join(org_directory, excel_file_name)
        workbook = openpyxl.load_workbook(excel_file_path_m)
        sheet = workbook.active
        rebuild(driver, wait, excel_file_path, org_excel,ws = sheet,wb=workbook, excel_file_path_m=excel_file_path_m)
    else:
        node_amount = init(driver, wait, org_directory)
        if node_amount == False:
            return
        driver.refresh()
    
        os.makedirs(org_directory, exist_ok=True)
        #设计党组织基本信息表头
        columns_base_info = ["序号","党组织全称", "组织树", "党组织简称", "党内统计用党组织简称", "成立日期", "党组织编码", "党组织联系人", "联系电话", "组织类别", 
            '是否具有"审批预备党员权限"', "功能型党组织", "党组织所在单位情况", "党组织所在行政区划", "批准成立的上级党组织", "是否为新业态", "驻外情况", 
            "党组织曾用名", "单位名称（全称）", "UUID", "有无统一社会信用代码", "法人单位统一社会信用代码", "单位性质类别", 
            "法人单位标识", "建立党组情况", "法人单位建立党组织情况", "在岗职工人数", "企业控制（控股）情况", "企业规模", "单位所在目录", "单位隶属关系", "单位所在行政区划", "单位名称(全称)", "机构类型", "法人单位统一社会信用代码"
            , "新经济行业", "经济行业", "经济类型", "新经济类型", "成立日期", "注册地行政区划", "注册地址", "组织机构代码", "上级主管部门名称", "单位隶属关系"]


        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sheet1"
        for i, value in enumerate(columns_base_info, start=1):
            ws.cell(row=1, column = i, value=value)
        wb.save(excel_file_path)
        #使用openpyxl库来加载一个已存在的Excel工作簿
        org_excel = load_workbook(excel_file_path)
        #打印调试信息
        print(f"文件 '{excel_file_path}' 已成功创建。")
        #启动同步
        synchronizing(wait, org_excel, excel_file_path, driver, ws, wb, node_amount)


def synchronizing(wait, org_excel, org_excel_path, driver, ws, wb, node_amount):
    org_totol_amount = node_amount
    config = configparser.ConfigParser()
    config.read('config.ini')
    org_directory = config.get('Paths_org_info', 'org_info_path')
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党组织名单.xlsx"
    excel_file_path_m = os.path.join(org_directory, excel_file_name)
    workbook = openpyxl.load_workbook(excel_file_path_m)
    sheet = workbook.active
    
    for i in range(1, org_totol_amount+1):
        if amount_that_complete%20 == 0:
            driver.close()
            try:
                for handle in driver.window_handles:
                    driver.switch_to.window(handle)
                    if '管理员工作台' in driver.title:
                        break
            except:
                time.sleep(1)
            access_org_database(driver, wait)
            switch_role(wait)
            switch_item_org(wait)
            time.sleep(3)
        #！！！！！@
        
        string = sheet.cell(row = i, column = 3).value
        array = string.split('/')
        array = array[1:]
        #！！！！！@  从excel中读取
        current_node_text = array[-1]
        container = wait_return_subelement_absolute(wait, 1, xpath="//div[@class = 'tree_wrapper_div']")
        for i, node_value in enumerate(array):
            complete_path = wait_return_subelement_relative(1, container, xpath=f".//span[contains(text(), '{node_value}')]/../../..")
            arrow = wait_return_subelement_relative(1, complete_path, xpath="./span")
            html = complete_path.get_attribute("outerHTML")
            soup_html = BeautifulSoup(html, 'html.parser')
            span_element = soup_html.find('span')
            if span_element and 'expanded' in span_element.get('class'):
                continue
            else:
                scroll = wait_return_subelement_absolute(wait, 1, "//div[@class = 'tree_wrapper']")
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scroll)
                arrow.click()         #！！！！！！！！！！！！！！！！！！！！！！！！！！！！
        target = wait_return_subelement_relative_v2(1, container, xpath=f".//span[contains(text(), '{current_node_text}')]/..")
        target.click()
        downloading(file = org_excel, wait = wait, driver = driver, path = org_excel_path, rebuild=False, ws=ws, wb=wb, )
       
    
    
    #递归遍历树状列表完成每一个党组织的信息采集。
######    #点击根组织(//div[@class = "tree_wrapper"]//div[@role = "treeitem"]/div[@class = "fs-tree-node__content"])[1]
######    while (1):
######        try:
######            root_org = wait.until(EC.element_to_be_clickable((By.XPATH, "(//div[@class = 'tree_wrapper']//div[@role = 'treeitem']/div[@class = 'fs-tree-node__content'])[1]")))
######            break
######        except:
######            time.sleep(0.5)
######    while(1):
######        try:
######            root_org.click()
######            break
######        except:
######            root_org = wait.until(EC.element_to_be_clickable((By.XPATH, "(//div[@class = 'tree_wrapper']//div[@role = 'treeitem']/div[@class = 'fs-tree-node__content'])[1]")))
######            time.sleep(0.5)
######    #采集根组织信息
######    downloading(file = org_excel, wait = wait, driver = driver, path = org_excel_path)
######

    #获取根节点结构体//div[@class = "tre
    # 
    # e_wrapper"]//div[@role = "treeitem"]/div[@role = "group"]到tree_root
    #####while(1):
    #####    try:
    #####        tree_root = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class = 'tree_wrapper']//div[@role = 'treeitem']/div[@role = 'group']")))
    #####        break
    #####    except:
    #####        time.sleep(0.5)


    #item_html = tree_root.get_attribute('outerHTML')
    #with open('item_html.txt', 'w', encoding = 'utf-8') as file0:
        #file0.write(item_html)

    #recursion(tree_root)
####    recursion(tree_root = tree_root, file = org_excel, wait = wait, driver = driver, path = org_excel_path)
####
####
####
####def recursion(tree_root, file, wait, driver, path):
####    org_items = []
####    org_items = tree_root.find_elements(By.XPATH, ".//div[@role = 'treeitem']")
####    
####    #item_html0 = org_items[0].get_attribute('outerHTML')
####    #with open('item_html0.txt', 'w', encoding = 'utf-8') as file0:
####        #file0.write(item_html0)
####
####    #for 每个元素 in tree_root
####    for item in org_items:
####        # 每个元素.click()
####        item.click()
####        # 采集党组织信息
####        downloading(file, wait, driver, path)
####        # 查看元素下面是否有//span[@class = "is-leaf fs-tree-node__expand-icon fs-icon-caret-right"]
####        ##在这里检查线程关闭信号
####
####        if stop_event.is_set():
####            break
####        item_html = item.get_attribute('outerHTML')
####        soup = BeautifulSoup(item_html, 'html.parser')
####        first_child = soup.find()
####        first_grandchild = first_child.find() if first_child else None
####        leaf = first_grandchild.find('span', class_= "is-leaf fs-tree-node__expand-icon fs-icon-caret-right")
####        #如果有：
####        if leaf:
####            #next
####            next
####        #如果没有：
####        else:
####            #断言元素下面有//span[@class = "fs-tree-node__expand-icon fs-icon-caret-right"]
####                #断言异常
####            #点击//span[@class = "fs-tree-node__expand-icon fs-icon-caret-right"]
####            while (1):
####                try:
####                    arrow_down = item.find_element(By.XPATH, "./div[@class='fs-tree-node__content']/span[@class='fs-tree-node__expand-icon fs-icon-caret-right']")
####                    break
####                except:
####                    time.sleep(0.5)
####            while(1):
####                try:
####                    arrow_down.click()
####                    break
####                except:
####                    time.sleep(0.5)
####                    arrow_down = item.find_element(By.XPATH, "./div[@class='fs-tree-node__content']/span[@class='fs-tree-node__expand-icon fs-icon-caret-right']")
####            #元素下面的//div[@role = "group"]结构体作为新的根节点结构体new_tree_root
####            while(1):
####                try:
####                    new_tree_root = item.find_element(By.XPATH, ".//div[@role = 'group']")
####                    break
####                except:
####                    time.sleep(0.5)
####            #递归recursion(new_tree_root)
####            recursion(new_tree_root, file, wait, driver, path)

def downloading(file, wait, driver, path, rebuild, ws, wb):
    global amount_that_complete
    if rebuild == True:
        count = amount_that_complete
        countx = count + 1
    else:
        count = amount_that_complete
        count = count + 1
        countx = count + 1
    #填写序号#
    file.active.cell(row=countx, column=1).value = count
    ws.cell(row=countx, column = 1, value=count)
    
    #填写党组织全称#
    name_temp = wait.until(EC.visibility_of_element_located((By.XPATH, "//span[contains(text(), '党组织全称')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 2, value=name_temp)
    
    #循环断言
    while(1):
        try:
            df = ws.cell(row=countx, column = 2).value
            assert df != ""
            break
        except AssertionError:
            time.sleep(0.5)
            name_temp = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '党组织全称')]/../following-sibling::*/div[1]"))).text
            ws.cell(row=countx, column = 2, value=name_temp)
            
    #组织树 #wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[3]"))).text
    ws.cell(row=countx, column = 3, value="-")
    #党组织简称#
    si = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '党组织简称')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 4, value=si) 
    #党内统计用党组织简称#
    wu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '党内统计用党组织简称')]/following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 5, value=wu)
    #成立日期#
    liu = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),  '成立日期')]/../following-sibling::div/div[1]"))).text
    ws.cell(row=countx, column = 6, value=liu)
    #党组织编码#
    qi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '党组织编码')]/following-sibling::*/span"))).text
    ws.cell(row=countx, column = 7, value=qi)
    #党组织联系人#
    ba = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '党组织联系人')]/following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 8, value=ba)
    #联系电话#
    jiu = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '联系电话')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 9, value=jiu)
    #组织类别#
    shi = file.active.cell(row=countx, column=10).value = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '组织类别')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 10, value=shi)
    #是否具有"审批预备党员权限"#
    if "委员会" in shi:
        shiyi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '审批预备党员权限')]/following-sibling::div/span"))).text
        ws.cell(row=countx, column = 11, value=shiyi)
    elif "总支" in shi:
        ws.cell(row=countx, column = 11, value="-")
    else:
        ws.cell(row=countx, column = 11, value="-")
    #功能型党组织#
    shier= wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '功能型党组织')]/following-sibling::*/span"))).text
    ws.cell(row=countx, column = 12, value=shier)
    #党组织所在单位情况#
    shisan = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '党组织所在单位情况')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 13, value=shisan)
    #党组织所在行政区划#
    shisi = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(), '党组织所在行政区划')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 14, value=shisi)
    #批准成立的上级党组织#
    shiwu = wait.until(EC.presence_of_element_located((By.XPATH, "//span[contains(text(),  '批准成立的上级党组织')]/../following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 15, value=shiwu)
    #是否为新业态#
    shiliu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '是否为新业态')]/following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 16, value=shiliu)
    #驻外情况#
    shiqi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(),  '驻外情况')]/following-sibling::*/div[1]"))).text
    ws.cell(row=countx, column = 17, value=shiqi)
    #党组织曾用名，这里需要用soup判断是否有这一条
    html0 = driver.find_element(By.XPATH, '//div[contains(text(), "党组织曾用名")]/../..')
    _html0 = html0.get_attribute("style")
    if "display: none;" in _html0:
    #党组织曾用名，这里需要用soup判断是否有这一条
        ws.cell(row=countx, column = 18, value="无")
    else:
        shiba = wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),  '党组织曾用名')]/following-sibling::*//tbody/tr/td[2]/div/div"))).text
        ws.cell(row=countx, column = 18, value=shiba)

    #切换选项卡
    while(1):
        try:
            company_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '党组织单位信息')]")))
            break
        except:
            time.sleep(0.1)
    while(1):
        try:
            company_info.click()
            break
        except:
            company_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '党组织单位信息')]")))

    #单位名称（全称）#
    shijiu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '单位名称（全称）')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 19, value=shijiu)
    i_1 = 0
    while(i_1 < 15):
        try:
            df =  ws.cell(row=countx, column=19).value
            #print("df:",df)
            assert df != ""
            break
        except AssertionError:
            time.sleep(0.5)
            shijiu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '单位名称（全称）')]/following-sibling::div"))).text
            ws.cell(row=countx, column = 19, value=shijiu)
            i_1 = i_1 + 1

    #UUID#
    ershi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), 'UUID')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 20, value=ershi)
    #有无统一社会信用代码#
    ershiyi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '有无统一社会信用代码')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 21, value=ershiyi)
    #法人单位统一社会信用代码#
    ershier = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '法人单位统一社会信用代码')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 22, value=ershier)
    #单位性质类别#
    org_type = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '单位性质类别')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 23, value=org_type)
    #法人单位标识#
    ershisi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '法人单位标识')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 24, value=ershisi)
    #建立党组情况!
    temp_1 = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'info-content-view'])[2]")))
    temp_1_html = temp_1.get_attribute("outerHTML")
    if '建立党组情况' in temp_1_html:
        ershiwu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '建立党组情况')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 25, value=ershiwu)
    else:
        ershiwu = "-"
        ws.cell(row=countx, column = 25, value=ershiwu)
    #法人单位建立党组织情况!
    if '法人单位建立党组织情况' in temp_1_html:
        ershiliu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '法人单位建立党组织情况')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 26, value=ershiliu)
    else:
        ershiliu = "-"
        ws.cell(row=countx, column = 26, value=ershiliu)
    #在岗职工数#
    #如果字符串org_type里面包含字眼"公司"则执行下面两行代码
    if "公司" in org_type:
        ershiqi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '在岗职工数')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 27, value=ershiqi)
    #否则执行
    else:
        ershiqi = "-"
        ws.cell(row=countx, column = 27, value=ershiqi)
    #在企业控制（控股）情况
    if "公司" in org_type:
        ershiba = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '企业控制（控股）情况')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 28, value=ershiba)
    else:
        ershiba = "-"
        ws.cell(row=countx, column = 28, value=ershiba)
    #企业规模
    if "公司" in org_type:
        ershijiu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '企业规模')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 29, value=ershijiu)
    else:
        ershijiu = "-"
        ws.cell(row=countx, column = 29, value=ershijiu)
    #单位所在目录
    sanshi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '单位所在目录')]/following-sibling::div"))).text
    ws.cell(row=countx, column = 30, value=sanshi)
    #如果单位性质类别位行政机关，则补充单位隶属关系
    if "机关" in org_type:
        sishiwu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '隶属关系')]/following-sibling::div/div/span"))).text
        ws.cell(row=countx, column = 45, value=sishiwu)
    else:
        sishiwu = "-"
        ws.cell(row=countx, column = 45, value=sishiwu)

    #民营科技企业标识
    if "公司" in org_type:
        sanshiyi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '民营科技企业标识')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 31, value=sanshiyi)
    else:
        sanshiyi = "-"
        ws.cell(row=countx, column = 31, value=sanshiyi)
    # 单位所在行政区划
    sanshier = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '单位所在行政区划')]/following-sibling::div//input"))).text
    ws.cell(row=countx, column = 32, value=sanshier)
    # 判断参考信息（省标院等单位）是否存在
    while 1:
        try:
            readonly = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'read-only'])[1]")))
            break
        except:
            time.sleep(0.5)
    
    readonly_html = readonly.get_attribute('outerHTML')
    soup = BeautifulSoup(readonly_html, 'html.parser')
    
    if soup.find_all(string= lambda text: '机构类型' in text):
        # 单位名称(全称)
        sanshisan = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '单位名称(全称)')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 33, value=sanshisan)
        # 机构类型
        sanshisi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '机构类型')]/following-sibling::div//span[@style = 'margin-top: auto; margin-bottom: auto;']"))).text
        ws.cell(row=countx, column = 34, value=sanshisi)
        # 法人单位统一社会信用代码
        sanshiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//label[contains(text(), '法人单位统一社会信用代码')])[2]/following-sibling::div"))).text
        ws.cell(row=countx, column = 35, value=sanshiwu)
        # 新经济行业
        sanshiliu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '新经济行业')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 36, value=sanshiliu)
        # 经济行业
        sanshiqi = wait.until(EC.presence_of_element_located((By.XPATH, "(//label[contains(text(), '经济行业')]/following-sibling::div)[2]"))).text
        ws.cell(row=countx, column = 37, value=sanshiqi)
        # 经济类型
        sanshiba = wait.until(EC.presence_of_element_located((By.XPATH, "(//label[contains(text(), '经济类型')]/following-sibling::div)[1]//span[@style = 'margin-top: auto; margin-bottom: auto;']"))).text
        ws.cell(row=countx, column = 38, value=sanshiba)
        # 新经济类型
        sanshijiu = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '新经济类型')]/following-sibling::div//span[@style = 'margin-top: auto; margin-bottom: auto;']"))).text
        ws.cell(row=countx, column = 39, value=sanshijiu)
        # 成立日期
        sishi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '成立日期')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 40, value=sishi)
        # 注册地行政区划
        sishiyi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '注册地行政区划')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 41, value=sishiyi)
        # 注册地址
        sishier = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '注册地址')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 42, value=sishier)
        # 组织机构代码
        sishisan = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '组织机构代码')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 43, value=sishisan)
        # 上级主管部门名称
        sishisi = wait.until(EC.presence_of_element_located((By.XPATH, "//label[contains(text(), '上级主管部门名称')]/following-sibling::div"))).text
        ws.cell(row=countx, column = 44, value=sishisi)
        wb.save(path)

    else:
        sanshisan = "-"
        ws.cell(row=countx, column = 33, value=sanshisan)
        # 机构类型
        sanshisi = "-"
        ws.cell(row=countx, column = 34, value=sanshisi)
        # 法人单位统一社会信用代码
        sanshiwu = "-"
        ws.cell(row=countx, column = 35, value=sanshiwu)
        # 新经济行业
        sanshiliu = "-"
        ws.cell(row=countx, column = 36, value=sanshiliu)
        # 经济行业
        sanshiqi = "-"
        ws.cell(row=countx, column = 37, value=sanshiqi)
        # 经济类型
        sanshiba = "-"
        ws.cell(row=countx, column = 38, value=sanshiba)
        # 新经济类型
        sanshijiu = "-"
        ws.cell(row=countx, column = 39, value=sanshijiu)
        # 成立日期
        sishi = "-"
        ws.cell(row=countx, column = 40, value=sishi)
        # 注册地行政区划
        sishiyi = "-"
        ws.cell(row=countx, column = 41, value=sishiyi)
        # 注册地址
        sishier = "-"
        ws.cell(row=countx, column = 42, value=sishier)
        # 组织机构代码
        sishisan = "-"
        ws.cell(row=countx, column = 43, value=sishisan)
        # 上级主管部门名称
        sishisi = "-"
        ws.cell(row=countx, column = 44, value=sishisi)
        wb.save(path)
    # 切换选项卡到班子成员
    while 1:
        try:
            councilcard = wait.until(EC.element_to_be_clickable((By.ID, "tab-class")))
            break
        except:
            time.sleep(0.5)
    while 1:
        try:
            councilcard.click()
            break
        except:
            councilcard = wait.until(EC.element_to_be_clickable((By.ID, "tab-class")))
            time.sleep(0.5)
    # 采集班子成员信息
    table_council(name_temp, wait, driver)

    # right_arrow = wait_return_subelement_absolute_notmust(wait, 1, "//i[@class = 'fs-icon-arrow-right']", 3)
    
    # if right_arrow is not None:
    #     right_arrow.click()

    # 切换选项卡到惩戒信息
    while 1:
        try:
            RewaridsAndPunishments = wait.until(EC.element_to_be_clickable((By.ID, "tab-rewardsPunishments")))
            break
        except:
            time.sleep(0.5)
    while 1:
        try:
            RewaridsAndPunishments.click()
            break
        except:
            RewaridsAndPunishments = wait.until(EC.element_to_be_clickable((By.ID, "tab-rewardsPunishments")))
            time.sleep(0.5)
    # 采集奖励惩戒信息
    table_reward_punish(name_temp, wait)
    
    print("填写第",count,"个党组织",name_temp,"信息成功") 
    amount_that_complete = amount_that_complete + 1




def table_council(org_name: str, wait, driver):
    #获取数据存储目录
    global org_directory
    #获取当天日期
    today_date = datetime.now().strftime("%Y-%m-%d")
    #构建文件名称
    excel_file_name = f"{today_date}{org_name}班子成员信息库.xlsx"
    #构建完整路径
    excel_file_path = os.path.join(org_directory, excel_file_name)
    #创建一个dataframe表头为columns_base_info中的元素
    df = pd.DataFrame()
    #dataframe导出到excel
    df.to_excel(excel_file_path, index = False)
    #openpyexcel打开excel_file_path
    book = load_workbook(excel_file_path)
    sheet = book.active
    #将党组织的全称存放在单元格A1
    sheet['A1'] = org_name
    book.save(excel_file_path)
    #设计党组织委员会信息表头
    table_head = ["序号", "党内职务", "姓名", "公民身份证号码", "性别", "出生日期", "学历", "领导职务", "任职日期", "离职日期", "排序", "公司职务"]
    for i, value in enumerate(table_head, start = 1):
        sheet.cell(row = 2, column = i, value = value)
        book.save(excel_file_path)
    #读取表格容器保存在变量container里
    while(1):
        try:
            tbody = wait_return_subelement_absolute(wait, 1, xpath="(//span[contains(text(), '班子信息集')]/../../../div[3]//tbody)[1]")
            # tbody = wait.until(EC.visibility_of_element_located((By.XPATH, "(//span[contains(text(), '班子信息集')]/../../../div[3]//tbody)[1]")))
                                                                           #(//div[@class = 'fs-table__fixed-body-wrapper'])[3]//tbody
                                                                           #//div[@class = 'fs-table fs-table--fit fs-table--border fs-table--enable-row-transition fs-table--mini']/div[@class = 'fs-table__body-wrapper is-scrolling-none']//tbody
                                                                           #(//div[@class = 'fs-table__body-wrapper is-scrolling-left'])[5]//tbody
            break
        except:
            while 1:
                try:
                    councilcard = wait.until(EC.element_to_be_clickable((By.ID, "tab-class")))
                    break
                except:
                    time.sleep(0.5)
            while 1:
                try:
                    councilcard.click()
                    break
                except:
                    councilcard = wait.until(EC.element_to_be_clickable((By.ID, "tab-class")))
                    time.sleep(0.5)
            erro_html = driver.execute_script("return document.documentElement.outerHTML")
            with open('erro.txt', 'w', encoding = "utf-8") as file1:
                file1.write(erro_html)
            break

    #用beautifulsoup分析container，将表头保存在新的容器container_header里，将表体保存在新的容器container_body里
   # while(1):
        #try:
            #soup = BeautifulSoup(tbody, 'html.parser')
            #break
        #except TypeError:
            #time.sleep(0.5)
            #tbody = wait.until(EC.visibility_of_element_located((By.XPATH, "(//div[@class = 'fs-table__header-wrapper']/following-sibling::div[@class = 'fs-table__body-wrapper is-scrolling-left'])[2]//tbody")))
    soup = BeautifulSoup(tbody.get_attribute("outerHTML"), 'html.parser')
    tr = soup.find_all('tr')


    #从tr里提取数据保存在自单元格A3始向右的区域内
    for j, every_row in enumerate(tr, start = 1):
        # 从elm_tr里解析出td, 保存在td数组里
        col = every_row.find_all('td')     
        #将序号i填入第j+2行，第1列
        sheet.cell(row = j+2, column = 1, value = j)
        book.save(excel_file_path)

        #将党内职务填入第j+2行，第2列
        position = col[1].find_all('span')[-1]
        position_content = position.get_text(strip = True)
        sheet.cell(row = j+2, column = 2, value = position_content)
        book.save(excel_file_path)

        #将姓名填入第j+2行，第3列
        name = col[2].find_all('span')[-1]
        name_content = name.get_text(strip = True)
        sheet.cell(row = j+2, column = 3, value = name_content)
        book.save(excel_file_path)

        #将公民身份证号码职务填入第j+2行，第4列
        idcard_no = col[3].find_all('div')[-1]
        idcard_no_content = idcard_no.get_text(strip = True)
        sheet.cell(row = j+2, column = 4, value = idcard_no_content)
        book.save(excel_file_path)

        #将性别填入第j+2行，第5列
        gender = col[4].find_all('span')[-1]
        gender_content = gender.get_text(strip = True)
        sheet.cell(row = j+2, column = 5, value = gender_content)
        book.save(excel_file_path)

        #将出生日期填入第j+2行，第6列
        birthday = col[5].find_all('span')[-1]
        birthday_content = birthday.get_text(strip = True)
        sheet.cell(row = j+2, column = 6, value = birthday_content)
        book.save(excel_file_path)

        #将学历填入第j+2行，第7列
        edu_qual = col[6].find_all('span')[-1]
        edu_qual_content = edu_qual.get_text(strip = True)
        sheet.cell(row = j+2, column = 7, value = edu_qual_content)
        book.save(excel_file_path)

        #将领导职务填入第j+2行，第8列span
        leader_position = col[7].find_all('span')[-1]
        leader_position_content = leader_position.get_text(strip = True)
        sheet.cell(row = j+2, column = 8, value = leader_position_content)
        book.save(excel_file_path)

        #将任职日期填入第j+2行，第9列span
        appointmentdate = col[8].find_all('span')[-1]
        appointmentdate_content = appointmentdate.get_text(strip = True)
        sheet.cell(row = j+2, column = 9, value = appointmentdate_content)
        book.save(excel_file_path)

        #将离职日期填入第j+2行，第10列span
        resignationdate = col[9].find_all('span')[-1]
        resignationdate_content = resignationdate.get_text(strip = True)
        sheet.cell(row = j+2, column = 10, value = resignationdate_content)
        book.save(excel_file_path)

        #将排序填入第j+2行，第11列div
        sort = col[10].find_all('div')[-1]
        sort_content = sort.get_text(strip = True)
        sheet.cell(row = j+2, column = 11, value = sort_content)
        book.save(excel_file_path)

        #将公司职务填入第j+2行，第12列div
        companyposition = col[11].find_all('div')[-1]
        companyposition_content = companyposition.get_text(strip = True)
        sheet.cell(row = j+2, column = 12, value = companyposition_content)
        book.save(excel_file_path)
    #释放资源
    book.close()
    


def table_reward_punish(org_name: str, wait):
    #获取数据存储目录
    global org_directory
    #获取当天日期
    today_date = datetime.now().strftime("%Y-%m-%d")
    #构建文件名称
    excel_file_name = f"{today_date}{org_name}奖惩信息.xlsx"
    #构建完整路径
    excel_file_path = os.path.join(org_directory, excel_file_name)
    #创建一个dataframe表头为columns_base_info中的元素
    df = pd.DataFrame()
    #dataframe导出到excel
    df.to_excel(excel_file_path, index = False)
    #openpyexcel打开excel_file_path
    book = load_workbook(excel_file_path)
    sheet = book.active
    #将党组织的全称存放在单元格A1
    sheet['A1'] = org_name
    book.save(excel_file_path)
    #设计党组织委员会信息表头
    table_head = ["序号", "奖惩名称", "批准机关", "批准日期"]
    #将设计好的表头填入excel
    for i, value in enumerate(table_head, start = 1):
        sheet.cell(row = 2, column = i, value = value)
        book.save(excel_file_path)
    #寻找奖惩块
    while 1: 
        try:
            box_table = wait.until(EC.visibility_of_element_located((By.ID, "pane-rewardsPunishments")))
            break        
        except:
            time.sleep(0.5)
    #soup分析奖惩块
    table_html = box_table.get_attribute('outerHTML')
    soup = BeautifulSoup(table_html, 'html.parser')
    #检查是否有数据，如果没有则全填-
    if soup.find_all(string = lambda text: '暂无数据' in text):
        for cell in sheet['A3:D3'][0]:
            cell.value = '-'
            book.save(excel_file_path)
    #如果有数据则填表
    else:
        #寻找表体tbody
        while 1:
            try:
                tbody = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class = 'fs-table__body-wrapper is-scrolling-none']//tbody")))
                break
            except:
                time.sleep(0.5)
        #soup分析tbody
        tbody_html = tbody.get_attribute("outerHTML")
        soup = BeautifulSoup(tbody_html, 'html.parser')
        #soup提取所有行
        rows= soup.find_all('tr')
        
        #对于每一行
        for j, each_row in enumerate(rows, start = 1):
            #soup提取所有列
            col = each_row.find_all('td')

            #将序号j填入第j+2行，第1列
            sheet.cell(row = j+2, column = 1, value = j)
            book.save(excel_file_path)

            #将奖惩名称填入第j+2行，第2列
            try:
                name = col[0].find_all('div')[-1]
            except:
                name = col[0].find_all('span')[-1]
            name_content = name.get_text(strip = True)
            sheet.cell(row = j+2, column = 2, value = name_content)
            book.save(excel_file_path)

            #将批准机关填入第j+2行，第3列
            try:
                justice = col[1].find_all('span')[-1]
            except:
                justice = col[1].find_all('div')[-1]
            justice_content = justice.get_text(strip = True)
            sheet.cell(row = j+2, column = 3, value = justice_content)
            book.save(excel_file_path)

            #将批准日期填入第j+2行，第4列
            try:
                passdate = col[2].find_all('span')[-1]
            except:
                passdate = col[2].find_all('div')[-1]
            passdate_content = passdate.get_text(strip = True)
            sheet.cell(row = j+2, column = 4, value = passdate_content)
            book.save(excel_file_path)
    #释放资源
    book.close()


def element_exists(driver, by, value):
    try:
        driver.find_element(by, value)
        return True
    except NoSuchElementException:
        return False
    

def access_info_page(wait, row):
    xpath = f"(//table[@class='fs-table__body'])[3]/tbody/tr[{row}]/td[3]"
    while 1:
        try:
            org = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            break
        except:
            time.sleep(0.1)
    while 1:
        try:
            org.click()
            break
        except:
                while 1:
                    try:
                        org = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                        break
                    except:
                        time.sleep(0.1)
    print(f"进入党员个人页面成功") 