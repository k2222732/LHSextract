#检查谁没上传主题党日、灯塔大课堂（仅党委有此功能）
import dev_func
from base.waitclick import *
from selenium.webdriver.support.ui import WebDriverWait
import login
import os
import pandas as pd
from tkinter import messagebox
import base.rwconfig as conf
import base.write_entry as cursor
from base.boot import *
from base.mystruct import TreeNode
import time
from base.solv_date import calculate_date_add
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
import dev_func
from bs4 import BeautifulSoup
import middle.switch_role
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from base.mystruct import TreeNode
from base.waitclick import *
from base.write_entry import *
from selenium.common.exceptions import NoSuchElementException
import re
from selenium.webdriver.common.keys import Keys
from datetime import datetime
from openpyxl import load_workbook
import openpyxl
import do_upload_dkt

ztdr = []
dkt = []
quanti = []
kssj = '2024-06'#开始时间
jssj = '2024-06'#结束时间

def main():
    global ztdr
    global dkt
    global quanti
    driver = login._main()
    wait = WebDriverWait(driver, 2, 0.5)
    #获取全体支部名单
    dev_func.access_dev_database(driver, wait)
    dev_func.switch_role(wait, driver)
    commen_button(wait, driver, xpath="(//span[contains(text(), '人员信息')])[1]")
    #(//span[contains(text(), '人员信息')])[1]
    org_tree = TreeNode()
    do_upload_dkt.synchronizing_org(wait, driver, input_node=org_tree)
    org_tree.copy_to_list(quanti)
    try:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if '管理员工作台' in driver.title:
                break
    except:
        time.sleep(1)
    middle.switch_role.access_e_shandong(driver, wait)
    middle.switch_role.switch_role(wait)
    commen_button(wait, driver, xpath="//span[contains(text(), '数据统计')]")
    commen_button(wait, driver, xpath="(//span[contains(text(), '支部活动情况')])[1]")
    while 1:
        try:
            driver.find_element(By.XPATH, "//div[@id = 'tab-3']")
            break
        except:
            time.sleep(1)
    commen_button(wait, driver, xpath="//span[contains(text(), '列表设置')]")

    commen_button(wait, driver, xpath="(//div[contains(text(), '其他学习内容')])[3]/../following-sibling::td//label")
    commen_button(wait, driver, xpath="(//div[contains(text(), '谈心谈话')])[3]/../following-sibling::td//label")
    commen_button(wait, driver, xpath="(//div[contains(text(), '民主评议党员')])[3]/../following-sibling::td//label")
    commen_button(wait, driver, xpath="(//div[contains(text(), '组织生活会')])[3]/../following-sibling::td//label")
    commen_button(wait, driver, xpath="(//div[contains(text(), '党课')])[3]/../following-sibling::td//label")
    commen_button(wait, driver, xpath="(//div[contains(text(), '党小组会')])[3]/../following-sibling::td//label")
    commen_button(wait, driver, xpath="(//table[@class = 'fs-table__body'])[3]//tr[7]/td//label")
    commen_button(wait, driver, xpath="(//table[@class = 'fs-table__body'])[3]//tr[6]/td//label")
    commen_button(wait, driver, xpath="(//table[@class = 'fs-table__body'])[3]//tr[5]/td//label")
    commen_button(wait, driver, xpath="(//table[@class = 'fs-table__body'])[3]//tr[3]/td//label")
    commen_button(wait, driver, xpath="(//table[@class = 'fs-table__body'])[3]//tr[2]/td//label")
    commen_button(wait, driver, xpath="(//table[@class = 'fs-table__body'])[3]//tr[1]/td//label")
    commen_button(wait, driver, xpath="//span[contains(text(), '确 定')]")
    #(//input[@placeholder = '请选择'])[1]
    input_text(wait, driver, xpath="(//input[@placeholder = '请选择'])[1]", text=kssj)
    #(//input[@placeholder = '请选择'])[2]
    input_text(wait, driver, xpath="(//input[@placeholder = '请选择'])[2]", text=jssj)
    commen_button(wait, driver, xpath="//span[contains(text(), '查询')]")
    # #同步主题党日支部
    commen_button(wait, driver, xpath="(//tbody[1]/tr[1]/td[2]//span)[1]")
    totalnum_ztdr = get_amountof_member(wait)
    synchronizing_ztdr(wait, totalnum_ztdr)
    commen_button(wait, driver, xpath="(//I[@class = 'fs-dialog__close fs-icon fs-icon-close'])[1]")
    #同步灯塔大课堂支部
    commen_button(wait, driver, xpath="(//tbody[1]/tr[1]/td[4]//span)[1]")
    totalnum_dkt = get_amountof_member(wait)
    synchronizing_dkt(wait, totalnum_dkt)
    commen_button(wait, driver, xpath="(//I[@class = 'fs-dialog__close fs-icon fs-icon-close'])[1]")
    

    # mfilepath  = os.path.abspath(os.getcwd())+"\三会一课未上传"
    # today_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # excel_file_name = f"{today_date}三会一课未上传.xlsx"
    # excel_file_path = os.path.join(mfilepath, excel_file_name)
    # 通过链接得到绝对路径
    # columns = ["未上传大课堂的企业","未上传主题党日的企业"]
    # df = pd.DataFrame(columns=columns)
    # df.to_excel(excel_file_path, index=False)
    # member_excel = load_workbook(excel_file_path)


    difference_dkt = [item for item in quanti if item not in dkt]
    difference_ztdr = [item for item in quanti if item not in ztdr]

    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"
    # 循环写入数组 a 的元素，从单元格 A2 开始
    for i, value in enumerate(quanti, start=2):
        ws[f'A{i}'] = value

    for i, value in enumerate(difference_dkt, start=2):
        ws[f'B{i}'] = value

    for i, value in enumerate(difference_ztdr, start=2):
        ws[f'C{i}'] = value
    # 保存 Excel 文件
        ws['A1'] = "全体党支部名单"
        ws['B1'] = "未上传大课堂的企业"
        ws['C1'] = "未上传主题党日的企业"
    excel_file_path = os.path.abspath(os.getcwd())+"\三会一课未上传\三会一课未上传名单.xlsx"
    wb.save(excel_file_path)


    #悬停
    input("")
    



def get_amountof_member(wait):
    time.sleep(1)
    char_amountof_member = wait.until(EC.presence_of_element_located((By.XPATH, "(//span[@class = 'fs-pagination__total'])[2]"))).text
    int_amountof_members = re.findall(r'\d+', char_amountof_member)
    int_amountof_member = int(int_amountof_members[0]) if int_amountof_members else None
    return int(int_amountof_member)


def synchronizing_ztdr(wait, member_total_amount):
    amount_that_complete = 0
    page_number_used = 1
    while amount_that_complete < member_total_amount:
        page_number = int(amount_that_complete / 30 + 1)
        row_number = int(amount_that_complete % 30 + 1)
        #time.sleep(0.1)
        input_page = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class = 'fs-input fs-input--small fs-pagination__editor is-in-pagination']/input")))
        input_page.send_keys(Keys.CONTROL + "a")
        input_page.send_keys(Keys.BACKSPACE)
        input_page.send_keys(page_number)
        input_page.send_keys(Keys.RETURN)
        
        if page_number - page_number_used != 0:
            time.sleep(1)
        access_info_page_ztdr(wait, row_number)
        amount_that_complete = amount_that_complete + 1
        page_number_used = page_number

        ##在这里检查线程关闭信号

def access_info_page_ztdr(wait, row_number):
    #f"(//table[@class = 'fs-table__body'])[2]//tbody/tr[{row_number}]/td[2]"
    global ztdr
    element = wait.until(EC.presence_of_element_located((By.XPATH, f"(//table[@class = 'fs-table__body'])[2]//tbody/tr[{row_number}]/td[2]")))
    element_name = element.get_attribute("textContent")
    ztdr.append(element_name)
    

def synchronizing_dkt(wait, member_total_amount):
    amount_that_complete = 0
    while amount_that_complete < member_total_amount:
        page_number = int(amount_that_complete / 30 + 1)
        row_number = int(amount_that_complete % 30 + 1)
        #time.sleep(0.1)
        input_page = wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@class = 'fs-input fs-input--small fs-pagination__editor is-in-pagination']/input")))
        input_page.send_keys(Keys.CONTROL + "a")
        input_page.send_keys(Keys.BACKSPACE)
        input_page.send_keys(page_number)
        input_page.send_keys(Keys.RETURN)
        access_info_page_dkt(wait, row_number)
        amount_that_complete = amount_that_complete + 1

        ##在这里检查线程关闭信号

def access_info_page_dkt(wait, row_number):
    #f"(//table[@class = 'fs-table__body'])[2]//tbody/tr[{row_number}]/td[2]"
    global dkt
    element = wait.until(EC.presence_of_element_located((By.XPATH, f"(//table[@class = 'fs-table__body'])[2]//tbody/tr[{row_number}]/td[2]")))
    element_name = element.get_attribute("textContent")
    dkt.append(element_name)

if __name__ == "__main__":
    main()