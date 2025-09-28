from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import time 
import os
import re
import configparser
import threading
from bs4 import BeautifulSoup
from base.waitclick import *
from base.write_entry import *
from middle.dev_download import *
from openpyxl import load_workbook
import openpyxl
from base.validate import *
stop_event = threading.Event()




def access_info_page(wait, rowx, file, path, temp_countx, wb, ws):
    #所属党组织在这一步录入
    lishudangzuzhi = wait_return_subelement_absolute(wait, time_w = 0.5, xpath = f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[4]//a").text
    #lishudangzuzhi = wait.until(EC.visibility_of_element_located((By.XPATH, f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[5]//a"))).text
    
    ws.cell(row=temp_countx+2, column=28, value = lishudangzuzhi)
    if (temp_countx)%100 == 0:
        wb.save(path)
    wait_click_xpath(wait, time_w = 0.5, xpath = f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[1]//a")
    name = wait_return_subelement_absolute(wait, time_w = 0.5, xpath = f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[1]//a").text
    if '一线入党' in name:
        return True
    else:
        return False


def switch_formal_mem(wait):
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    while 1:    
        try:
            time.sleep(1)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
            break
        except:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "正式党员")]')
    wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary']//span[contains(text(), '查询')]")
    while 1:
        time.sleep(0.5)
        t = wait_return_subelement_absolute(wait, time_w=0.5, xpath='(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
        u=t.get_attribute("value")
        if u =="正式党员":
            break
        else:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            while 1:    
                try:
                    time.sleep(1)
                    wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
                    break
                except:
                    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "正式党员")]')
            wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary is-plain']//span[contains(text(), '查询')]")


def switch_informal_mem(wait):
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    while 1:    
        try:
            time.sleep(1)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
            break
        except:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "预备党员")]')
    wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary']//span[contains(text(), '查询')]")
    while 1:
        time.sleep(0.5)
        t = wait_return_subelement_absolute(wait, time_w=0.5, xpath='(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
        u=t.get_attribute("value")
        if u =="预备党员":
            break
        else:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            while 1:    
                try:
                    time.sleep(1)
                    wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
                    break
                except:
                    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "预备党员")]')
            wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary is-plain']//span[contains(text(), '查询')]")

def switch_devtarg(wait):
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    while 1:    
        try:
            time.sleep(1)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
            break
        except:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "发展对象")]')
    wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary']//span[contains(text(), '查询')]")
    while 1:
        time.sleep(0.5)
        t = wait_return_subelement_absolute(wait, time_w=0.5, xpath='(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
        u=t.get_attribute("value")
        if u =="发展对象":
            break
        else:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            while 1:    
                try:
                    time.sleep(1)
                    wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
                    break
                except:
                    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "发展对象")]')
            wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary is-plain']//span[contains(text(), '查询')]")

def switch_activist(wait):
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    while 1:    
        try:
            time.sleep(1)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
            break
        except:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "入党积极分子")]')
    wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary']//span[contains(text(), '查询')]")
    while 1:
        time.sleep(0.5)
        t = wait_return_subelement_absolute(wait, time_w=0.5, xpath='(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
        u=t.get_attribute("value")
        if u =="入党积极分子":
            break
        else:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            while 1:    
                try:
                    time.sleep(1)
                    wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
                    break
                except:
                    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "入党积极分子")]')
            wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary is-plain']//span[contains(text(), '查询')]")

def switch_applicant(wait):
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    while 1:    
        try:
            time.sleep(1)
            wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
            break
        except:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
    wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "入党申请人")]')
    wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary']//span[contains(text(), '查询')]")
    while 1:
        time.sleep(0.5)
        t = wait_return_subelement_absolute(wait, time_w=0.5, xpath='(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
        u=t.get_attribute("value")
        if u =="入党申请人":
            break
        else:
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//*[contains(text(), "人员信息")]/..)[1]')
            wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            while 1:    
                try:
                    time.sleep(1)
                    wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'el-input el-input--suffix is-focus']")))
                    break
                except:
                    wait_click_xpath(wait, time_w = 0.5, xpath = '(//div[@class = "el-form-item__content"])[3]/div[@class = "el-select"]//input')
            wait_click_xpath(wait, time_w = 0.5, xpath = '//ul[@class = "el-scrollbar__view el-select-dropdown__list"]//li//span[contains(text(), "入党申请人")]')
            wait_click_xpath(wait, time_w = 0.5, xpath = "//button[@class = 'el-button el-button--primary is-plain']//span[contains(text(), '查询')]")


def set_amount_perpage(wait):
    wait_click_xpath(wait, time_w = 0.5, xpath = "//span[contains(text(),'前往')]/preceding-sibling::span[1]")
    wait_click_xpath(wait, time_w = 0.5, xpath = "//span[contains(text(), '100条/页')]")
    #num = get_total_amount_list(wait, xpath="(//input[@class = 'el-input__inner'])[4]")
    #while 1:
    #    if num==100:
    #        break
    #    else:
    #        time.sleep(1)
    #        wait_click_xpath(wait, time_w = 0.5, xpath = "//span[contains(text(),'前往')]/preceding-sibling::span[1]")
    #        wait_click_xpath(wait, time_w = 0.5, xpath = "//span[contains(text(), '100条/页')]")        

def get_total_amount_list(wait, xpath):
    time.sleep(1)
    try:
        can= wait_return_subelement_absolute(wait, time_w = 0.5, xpath=xpath)
        char_amountof_member =can.get_attribute("value")
        int_amountof_members = re.findall(r'\d+', char_amountof_member)
        int_amountof_member = int(int_amountof_members[0]) if int_amountof_members else None
        return int(int_amountof_member)
    except:
        #//div[@class = "el-table y_table el-table--fit el-table--border el-table--enable-row-hover el-table--enable-row-transition"]
        can = wait_return_subelement_absolute(wait, 0.5, "//main[@class ='el-main main']")
        can_html = can.get_attribute('outerHTML')
        soup = BeautifulSoup(can_html, 'html.parser')
        result = soup.find('span', class_ = "el-table__empty-text")
        if result:
            return 0
        else:
            return -1