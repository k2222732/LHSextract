from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import time 
import os
import re
import configparser
import threading
from bs4 import BeautifulSoup
from base.waitclick import *
from base.write_entry import *

def jibenxinxi_download(file, path, wait, countx, count, wb, ws):
    #填写序号
    yi = count
    
    ws.cell(row = countx, column = 1, value =yi)
    #姓名
    while 1:
        try:
            name_temp = wait.until(EC.visibility_of_element_located((By.XPATH, "(//td[contains(text(), '姓名')]//following-sibling::td)[1]"))).text
            break
        except:
            pass
    er = name_temp
    ws.cell(row = countx, column = 2, value =er)
    #循环断言
    while(1):
        try:
            df = ws.cell(row=countx, column=2).value
            assert df != ""
            break
        except AssertionError:
            time.sleep(0.5)
            er = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '姓名')]//following-sibling::td)[1]"))).text
            ws.cell(row = countx, column = 2, value =er)
    #性别
    san = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '性别')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 3, value =san)
    #公民身份证号码
    si = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '公民身份')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 4, value =si)
    #民族
    wu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '民族')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 5, value =wu)
    #出生日期
    liu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '出生日期')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 6, value =liu)
    #学历
    bitian_located(file, path, countx, column=7, wait = wait, str = "(//td[contains(text(), '学历')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #申请入党日期
    ba = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '入党申请书')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 8, value =ba)
    #手机号码
    jiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '联系')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 9, value =jiu)
    #背景信息
    bitian_located(file, path, countx, column=10, wait = wait, str = "(//td[contains(text(), '背景信息')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #工作岗位
    bitian_located(file, path, countx, column=11, wait = wait, str = "(//td[contains(text(), '工作岗位')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #政治面貌
    bitian_located(file, path, countx, column=12, wait = wait, str = "(//td[contains(text(), '政治面貌')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #接受申请党组织
    shisan = wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '接收入党申请的党组织')])[1]//following-sibling::td").get_attribute("textContent")
    ws.cell(row = countx, column = 13, value =shisan)
    return name_temp


def jijifenzi_download(path, file, wait, countx, control, wb, ws):
    #切换到积极分子信息
        while(1):
            try:
                company_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '入党积极分子基本信息')]")))
                break
            except:
                time.sleep(0.1)
        while(1):
            try:
                company_info.click()
                break
            except:
                company_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '入党积极分子基本信息')]")))
        #籍贯
        jiguan_try = 1
        while 1:      
            shisi = wait.until(EC.presence_of_element_located((By.XPATH, "(//table[@class = 'y_table form-table']//span[contains(text(),'籍贯')])[2]/../following-sibling::td"))).text
            if shisi == "" and jiguan_try < 5:
                jiguan_try += 1
                pass
            else:
                ws.cell(row = countx, column = 14, value =shisi)
                break
  
        
        shiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//span[contains(text(), '入团日期')])[2]/../following-sibling::td"))).text
        ws.cell(row = countx, column = 15, value =shiwu)
        #参加工作日期
        shiliu = wait.until(EC.presence_of_element_located((By.XPATH, "((//span[contains(text(), '参加工作日期')])[2]/../following-sibling::td)[1]"))).text
        ws.cell(row = countx, column = 16, value =shiliu)
        #申请入党日期
        shiqi = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '递交入党申请书日期')])[2]/following-sibling::td"))).text
        ws.cell(row = countx, column = 17, value =shiqi)
        #确定入党积极分子日期
        shiba = wait.until(EC.presence_of_element_located((By.XPATH, "(//span[contains(text(), '确定入党积极分子日期')])[1]/../following-sibling::td"))).text
        ws.cell(row = countx, column = 18, value =shiba)
        #工作单位及职务
        shijiu = wait.until(EC.presence_of_element_located((By.XPATH, "((//td[contains(text(), '工作单位及职务')])[2]//following-sibling::td)[1]"))).text
        
        ws.cell(row = countx, column = 19, value =shijiu)
        #家庭住址
        ershi = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '现居住地')])[2]/following-sibling::td[1]"))).text
        ws.cell(row = countx, column = 20, value =ershi)
        
        if control == 3 or 4:
            ershiyi = "-"
            ws.cell(row = countx, column = 21, value =ershiyi)

            ershisan = "-"
            ws.cell(row = countx, column = 23, value =ershisan)

            ershisi = "-"
            ws.cell(row = countx, column = 24, value =ershisi)

            ershiwu = "-"
            ws.cell(row = countx, column = 25, value =ershiwu)

            ws.cell(row=countx, column=26).value = ws.cell(row=countx, column=28).value
        
        elif control == 1:
            pass
        elif control == 5:
            pass

        # 人员类型
        if control == 3:
            ershiqi = "发展对象"
            ws.cell(row = countx, column = 27, value =ershiqi)
            
        elif control == 4:
            ershiqi = "积极分子"
            ws.cell(row = countx, column = 27, value =ershiqi)
        



def yubeidangyuan(path, file, wait, countx, control, wb, ws):
    #切换到预备党员
        while(1):
            try:
                company_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '预备党员')]")))
                break
            except:
                time.sleep(0.1)
        while(1):
            try:
                company_info.click()
                break
            except:
                company_info = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[contains(text(), '预备党员')]")))
        # 加入党组织日期
        time.sleep(2)
        ershiyi = wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '加入党组织日期')])[1]//following-sibling::td[1]").get_attribute("textContent")
        ws.cell(row = countx, column = 21, value =ershiyi)
        # 转为正式党员日期
        ershier = '-'#wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '转为正式党员日期')])[1]//following-sibling::td[1]").get_attribute("textContent")
        ws.cell(row = countx, column = 22, value =ershier)
        # 人员类别

        ershisan = wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '人员类别')])[1]//following-sibling::td[1]").get_attribute("textContent")
        try_num = 1
        while 1:
            if ershisan != '' or try_num > 5:
                ershisan = wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '人员类别')])[1]//following-sibling::td[1]").get_attribute("textContent")
                break
            else:
                try_num += 1
        ws.cell(row = countx, column = 23, value =ershisan)
        # 党籍状态
        ershisi = wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '党籍状态')])[1]//following-sibling::td[1]").get_attribute("textContent")
        ws.cell(row = countx, column = 24, value =ershisi)
        # 入党类型
        ershiwu = wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '入党类型')])[1]//following-sibling::td[1]").get_attribute("textContent")
        ws.cell(row = countx, column = 25, value =ershiwu)
        # 所在党支部
        ershiliu = '缺项'#wait_return_subelement_absolute(wait, time_w = 0.5, xpath="(//td[contains(text(), '所属党组织')])[2]//following-sibling::td[1]").get_attribute("textContent")
        ws.cell(row = countx, column = 26, value =ershiliu)
        # 人员类型
        if control == 1:
            ershiqi = "正式党员"
            ws.cell(row = countx, column = 27, value =ershiqi)
        elif control == 2:
            ershiqi = "预备党员"
            ws.cell(row = countx, column = 27, value =ershiqi)
        

def formal_yixian_download(file, path, wait, countx, count, wb, ws):
    #填写序号
    yi = count
    
    ws.cell(row = countx, column = 1, value =yi)
    #姓名
    name_temp = wait.until(EC.visibility_of_element_located((By.XPATH, "(//td[contains(text(), '姓名')]//following-sibling::td)[1]"))).text
    er = name_temp
    ws.cell(row = countx, column = 2, value =er)
    #循环断言
    while(1):
        try:
            df = ws.cell(row=countx, column=2).value
            assert df != ""
            break
        except AssertionError:
            time.sleep(0.5)
            er = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '姓名')]//following-sibling::td)[1]"))).text
            ws.cell(row = countx, column = 2, value =er)
    #性别
    san = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '性别')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 3, value =san)
    #公民身份证号码
    si = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '公民身份')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 4, value =si)
    #民族
    wu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '民族')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 5, value =wu)
    #出生日期
    liu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '出生日期')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 6, value =liu)
    #学历
    bitian_located(file, path, countx, column=7, wait = wait, str = "(//td[contains(text(), '学历')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #申请入党日期
    ba = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '入党申请书')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 8, value =ba)
    #手机号码
    jiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '联系')]//following-sibling::td)[1]"))).text
    ws.cell(row = countx, column = 9, value =jiu)
    #背景信息
    bitian_located(file, path, countx, column=10, wait = wait, str = "(//td[contains(text(), '背景信息')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #工作岗位
    bitian_located(file, path, countx, column=11, wait = wait, str = "(//td[contains(text(), '工作岗位')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #政治面貌
    shier = '-'
    ws.cell(row = countx, column = 12, value =shier)
    #bitian_located(file, path, countx, column=12, wait = wait, str = "(//td[contains(text(), '政治面貌')]//following-sibling::td)[1]", wb=wb, ws=ws)
    #接受申请党组织
    shisan = '-'
    ws.cell(row = countx, column = 13, value =shisan)
    #入团日期
    shiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '入团日期')])/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 15, value =shiwu)
    #参加工作日期
    shiliu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '参加工作日期')])/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 16, value =shiliu)
    #申请入党日期
    shiqi = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '递交入党申请书日期')])/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 17, value =shiqi)
    #确定入党积极分子日期
    shiba = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '确定为入党积极分子日期')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 18, value =shiba)
    #工作单位及职务
    shijiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '工作单位及职务')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 19, value =shijiu)
    #家庭住址
    ershi = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '现居住地')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 20, value =ershi)
    #加入党组织日期
    ershiyi = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '加入党组织日期')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 21, value =ershiyi)
    #转为正式党员日期
    ershier = '-'
    ws.cell(row = countx, column = 22, value =ershier)
    #人员类别
    ershisan = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '人员类别')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 23, value =ershisan)
    #党籍状态
    ershisi = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '党籍状态')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 24, value =ershisi)
    #入党类型
    ershiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '入党类型')])[1]/following-sibling::td[1]"))).text
    ws.cell(row = countx, column = 25, value =ershiwu)
    #所在党支部
    ershiliu = wait.until(EC.presence_of_element_located((By.XPATH, "(//td[contains(text(), '所属党组织')])[1]/following-sibling::td[1]"))).text
    ws.cell(row=countx, column=26, value = ershiliu)
    ershiqi = "正式党员"
    ws.cell(row = countx, column = 27, value =ershiqi)
    return name_temp


def bitian_located(file, path, row, column, wait, str, wb, ws):
    z = wait_return_subelement_absolute(wait, time_w=0.5, xpath=str).text
    ws.cell(row=row, column=column, value=z)
    #file.active.cell(row=row, column=column).value = wait.until(EC.presence_of_element_located((By.XPATH, str))).text
    start_time = time.time()
    while(1):
            elapsed_time = time.time() - start_time
            if elapsed_time > 5:
                break
            try:
                df = ws.cell(row=row, column=column).value
                assert df != ""
                break
            except AssertionError:
                z = wait.until(EC.presence_of_element_located((By.XPATH, str))).text
                ws.cell(row=row, column=column, value=z)
                