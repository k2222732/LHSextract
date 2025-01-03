from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import time 
import os
import re
import configparser
import threading
from bs4 import BeautifulSoup
from base.waitclick import *
from base.write_entry import *
from middle.dev_download import *
from openpyxl import load_workbook
import openpyxl
from base.validate import *
from middle.enter_person_infopage import enter_person_infopage
from middle.dev_middle import *
stop_event = threading.Event()



def stop_dev_thread():
    stop_event.set()



captcha = ""
amount_mem_complete = 0
amount_infomem_complete = 0
amount_activist_complete = 0
amount_devtar_complete = 0
amount_applicant_complete = 0



mem_total_amount = 0
infomem_total_amount = 0
activist_total_amount = 0
devtar_total_amount = 0
applicant_total_amount = 0

dev_directory = ""
driver0 = ""
temp_countx = 0

def access_dev_database(driver, wait):
    global driver0
    driver0 = driver
    wait_click_xpath(wait, time_w = 0.5, xpath = '(//img[contains(@src, "发展党员纪实公示系统.png")])[2]')
    try:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if '发展党员纪实公示系统' in driver.title:
                break
        print(f"进入发展党员纪实公示系统成功")
    except:
        time.sleep(1)


def switch_role_dev(wait, driver):
    while 1:
        try:
            wait_click_xpath(wait, time_w = 0.5, xpath = "//i[@class = 'el-icon-caret-bottom']")
            print(f"进入角色下拉列表成功")
            config_file = 'config.ini'
            config = configparser.ConfigParser()
            config.read(config_file)
            role_name_dev = config.get('role_dev_name', 'name_role_dev', fallback='')
            element = wait_return_subelement_absolute(wait, time_w = 0.5, xpath = f'//li[contains(text(), "{role_name_dev}")]')
            html = element.get_attribute('outerHTML')
            soup = BeautifulSoup(html, 'html.parser')
            li = soup.find('li')
            li_class = li.get('class')
            sum = ''
            for ele in li_class:
                sum = sum + ele + ' '
                
            sum = sum[:-1]

            if 'el-dropdown-menu__item is-disabled red' in sum:
                break
            else:
                element.click()
                time.sleep(2)
                print(f"切换角色成功")
                break
        except:
            time.sleep(2)

def switch_role_e(wait, driver):
    while 1:
        try:
            wait_click_xpath(wait, time_w = 0.5, xpath = "//i[@class = 'el-icon-caret-bottom']")
            print(f"进入角色下拉列表成功")
            config_file = 'config.ini'
            config = configparser.ConfigParser()
            config.read(config_file)
            role_name_dev = config.get('role_e_name', 'name_role_e', fallback='')
            element = wait_return_subelement_absolute(wait, time_w = 0.5, xpath = f'//li[contains(text(), "{role_name_dev}")]')
            html = element.get_attribute('outerHTML')
            soup = BeautifulSoup(html, 'html.parser')
            li = soup.find('li')
            li_class = li.get('class')
            sum = ''
            for ele in li_class:
                sum = sum + ele + ' '
                
            sum = sum[:-1]

            if 'el-dropdown-menu__item is-disabled red' in sum:
                break
            else:
                element.click()
                time.sleep(2)
                print(f"切换角色成功")
                break
        except:
            time.sleep(2)

    








def new_excel(wait):
    global mem_total_amount
    global infomem_total_amount
    global devtar_total_amount
    global activist_total_amount
    global applicant_total_amount
    global dev_directory
    config = configparser.ConfigParser()
    config.read('config.ini')
    dev_directory = config.get('Paths_dev_info', 'dev_info_path')
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党员发展纪实信息库.xlsx"
    excel_file_path = os.path.join(dev_directory, excel_file_name)
    
    
    #if os.path.isfile(excel_file_path):
        #member_excel = load_workbook(excel_file_path)
        #rebuild(excel_file_path, wait, member_total_amount, member_excel, excel_file_path)
    #else:
        #在当前文件夹新建一个文件夹命名为"database"在里面新建一个文件夹名为"database_member"
        #os.makedirs(dev_directory, exist_ok=True)
        #在database_member文件夹里新建一个excel文件命名为"今天的日期"&"党员信息库"
        #在excel里的第一行建立表头，分别是"姓名、性别、公民身份号码、民族、出生日期、学历、人员类别、学位、所在党支部、手机号码、入党日期
        #转正日期、党龄、党龄校正值、新社会阶层类型、工作岗位、从事专业技术职务、是否农民工、现居住地、户籍所在地、是否失联党员、是否流动党员、
        #党员注册时间、注册手机号、党员增加信息、党员减少信息、入党类型、转正情况、入党时所在支部、延长预备期时间
    
    
    if os.path.isfile(excel_file_path):
        member_excel = load_workbook(excel_file_path)
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        rebuild(excel_file_path, wait, member_excel, wb, ws)
    else:
        os.makedirs(dev_directory, exist_ok=True)
        columns = ["序号", "姓名","性别","公民身份证号码","民族","出生日期","学历","申请入党日期","手机号码","背景信息","申请入党阶段工作岗位",
                "政治面貌","接受申请党组织","籍贯","入团日期","参加工作日期","申请入党日期","确定入党积极分子日期",
                "工作单位及职务","家庭住址","加入党组织日期","转为正式党员日期","人员类别","党籍状态","入党类型","积极分子阶段一线情况", "人员类型", "隶属党组织",
                "预备党员阶段一线情况", "预备党员阶段工作岗位", "预备党员阶段所属行业", "申请入党阶段一线情况", "积极分子阶段工作岗位"]
            
        # df = pd.DataFrame(columns=columns)
        # df.to_excel(excel_file_path, index=False)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sheet1"
        for i, value in enumerate(columns, start=1):
            ws.cell(row=1, column = i, value=value)
        

        wb.save(excel_file_path)
        
        member_excel = load_workbook(excel_file_path)
        print(f"文件 '{excel_file_path}' 已成功创建。")
        switch_formal_mem(wait)
        mem_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
        synchronizing(wait, member_excel, excel_file_path, control=1, wb=wb, ws=ws)
        if stop_event.is_set():
            return
        switch_informal_mem(wait)
        infomem_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
        synchronizing(wait, member_excel, excel_file_path, control=2, wb=wb, ws=ws)
        if stop_event.is_set():
            return
        switch_devtarg(wait)
        temp = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'searchForm flexCell node_flex']/following-sibling::div")))
        temp_html = temp.get_attribute("outerHTML")
        if "暂无数据" in temp_html:
            pass
        else:
            devtar_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
            synchronizing(wait, member_excel, excel_file_path, control=3, wb=wb, ws=ws)
        if stop_event.is_set():
            return
        switch_activist(wait)
        activist_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
        synchronizing(wait, member_excel, excel_file_path, control=4, wb=wb, ws=ws)
        if stop_event.is_set():
            return
        switch_applicant(wait)
        applicant_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
        synchronizing(wait, member_excel, excel_file_path, control=5, wb=wb, ws=ws)
        if stop_event.is_set():
            return


def rebuild(excel_file_path, wait, member_excel,wb,ws):
    global mem_total_amount
    global infomem_total_amount
    global devtar_total_amount
    global activist_total_amount
    global applicant_total_amount
    global amount_mem_complete
    global amount_infomem_complete
    global amount_activist_complete
    global amount_devtar_complete
    global amount_applicant_complete
    global temp_countx

    switch_formal_mem(wait)
    mem_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")  
    switch_informal_mem(wait)
    infomem_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
    switch_devtarg(wait)
    temp = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'searchForm flexCell node_flex']/following-sibling::div")))
    temp_html = temp.get_attribute("outerHTML")
    if "暂无数据" in temp_html:
        devtar_total_amount = 0
    else:
        devtar_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
    switch_activist(wait)
    activist_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
    switch_applicant(wait)
    applicant_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
    #先改变全局变量
    a = init_complete_amount(excel_file_path)
    temp_countx = a
    if temp_countx <0:
        print("请删除损坏的excel然后重新运行程序")
        return
    if a <= mem_total_amount:
        amount_mem_complete = a
        synchronizing(wait, member_excel, excel_file_path, control = 1,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_informal_mem(wait)
        synchronizing(wait, member_excel, excel_file_path, control=2,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_devtarg(wait)
        temp = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'searchForm flexCell node_flex']/following-sibling::div")))
        temp_html = temp.get_attribute("outerHTML")
        if "暂无数据" in temp_html:
            pass
        else:
            devtar_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
        synchronizing(wait, member_excel, excel_file_path, control=3,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_activist(wait)
        synchronizing(wait, member_excel, excel_file_path, control=4,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_applicant(wait)
        synchronizing(wait, member_excel, excel_file_path, control=5,wb=wb,ws=ws)
        if stop_event.is_set():
            return
    elif a>mem_total_amount and a <= mem_total_amount + infomem_total_amount:
        amount_mem_complete = mem_total_amount
        amount_infomem_complete = a - mem_total_amount
        synchronizing(wait, member_excel, excel_file_path, control = 2,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_devtarg(wait)
        temp = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'searchForm flexCell node_flex']/following-sibling::div")))
        temp_html = temp.get_attribute("outerHTML")
        if "暂无数据" in temp_html:
            pass
        else:
            devtar_total_amount = get_total_amount(wait, xpath = "//span[@class = 'el-pagination__total']")
            synchronizing(wait, member_excel, excel_file_path, control=3,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_activist(wait)
        synchronizing(wait, member_excel, excel_file_path, control=4,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_applicant(wait)
        synchronizing(wait, member_excel, excel_file_path, control=5,wb=wb,ws=ws)
        if stop_event.is_set():
            return
    elif a>mem_total_amount + infomem_total_amount and a <= mem_total_amount + infomem_total_amount + devtar_total_amount:
        amount_mem_complete = mem_total_amount
        amount_infomem_complete = infomem_total_amount
        amount_devtar_complete = a - (mem_total_amount + infomem_total_amount)
        synchronizing(wait, member_excel, excel_file_path, control = 3,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_activist(wait)
        synchronizing(wait, member_excel, excel_file_path, control=4,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_applicant(wait)
        synchronizing(wait, member_excel, excel_file_path, control=5,wb=wb,ws=ws)
        if stop_event.is_set():
            return
    elif a>mem_total_amount + infomem_total_amount + devtar_total_amount and a <= mem_total_amount + infomem_total_amount + devtar_total_amount + activist_total_amount:
        amount_mem_complete = mem_total_amount
        amount_infomem_complete = infomem_total_amount
        amount_devtar_complete = devtar_total_amount
        amount_activist_complete = a - (mem_total_amount + infomem_total_amount + devtar_total_amount)
        synchronizing(wait, member_excel, excel_file_path, control = 4,wb=wb,ws=ws)
        if stop_event.is_set():
            return
        switch_applicant(wait)
        synchronizing(wait, member_excel, excel_file_path, control=5,wb=wb,ws=ws)
        if stop_event.is_set():
            return
    elif a>mem_total_amount + infomem_total_amount + devtar_total_amount + activist_total_amount and a <= mem_total_amount + infomem_total_amount + devtar_total_amount + activist_total_amount + applicant_total_amount :
        amount_mem_complete = mem_total_amount
        amount_infomem_complete = infomem_total_amount
        amount_devtar_complete = devtar_total_amount
        amount_activist_complete = activist_total_amount
        amount_applicant_complete = a - (mem_total_amount + infomem_total_amount + devtar_total_amount + activist_total_amount)
        synchronizing(wait, member_excel, excel_file_path, control = 5,wb=wb,ws=ws)
        if stop_event.is_set():
            return

def init_complete_amount(excel_file_path):
    df = pd.read_excel(excel_file_path, sheet_name=0)
    row_count = count_non_empty_rows(excel_file_path, sheet_name=0)
    amount_that_complete = row_count - 3
    return amount_that_complete

def count_non_empty_rows(excel_file_path, sheet_name=0):
    # 加载Excel工作簿
    workbook = load_workbook(filename=excel_file_path, data_only=True)
    
    # 获取工作表（可以通过名称或索引获取）
    if isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        sheet = workbook[sheet_name]
    
    non_empty_row_count = 0
    
    # 逐行检查是否有数据
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            non_empty_row_count += 1
    return non_empty_row_count




def synchronizing(wait, member_excel, member_excel_path, control, wb, ws):
    global amount_mem_complete
    global amount_infomem_complete
    global amount_devtar_complete
    global amount_activist_complete
    global amount_applicant_complete

    global mem_total_amount
    global infomem_total_amount
    global devtar_total_amount
    global activist_total_amount
    global applicant_total_amount


    if control == 1:
        schedule(complete = amount_mem_complete, total = mem_total_amount, xpath = "//input[@type = 'number']", wait = wait, member_excel = member_excel, member_excel_path = member_excel_path, control = control, wb=wb, ws=ws)
    elif control == 2:
        schedule(complete = amount_infomem_complete, total = infomem_total_amount, xpath = "//input[@type = 'number']", wait = wait, member_excel = member_excel, member_excel_path = member_excel_path, control = control, wb=wb, ws=ws)
    elif control == 3:
        schedule(complete = amount_devtar_complete, total = devtar_total_amount, xpath = "//input[@type = 'number']", wait = wait, member_excel = member_excel, member_excel_path = member_excel_path, control = control, wb=wb, ws=ws)
    elif control == 4:
        schedule(complete = amount_activist_complete, total = activist_total_amount, xpath = "//input[@type = 'number']", wait = wait, member_excel = member_excel, member_excel_path = member_excel_path, control = control, wb=wb, ws=ws)
    elif control == 5:
        schedule(complete = amount_applicant_complete, total = applicant_total_amount, xpath = "//input[@type = 'number']", wait = wait, member_excel = member_excel, member_excel_path = member_excel_path, control = control, wb=wb, ws=ws)
    
    wb.save(member_excel_path)

def get_total_amount(wait, xpath):
    time.sleep(1)
    try:
        can= wait_return_subelement_absolute(wait, time_w = 0.5, xpath=xpath)
        char_amountof_member =can.get_attribute("textContent")
        int_amountof_members = re.findall(r'\d+', char_amountof_member)
        int_amountof_member = int(int_amountof_members[0]) if int_amountof_members else None
        return int(int_amountof_member)
    except:
        #//div[@class = "el-table y_table el-table--fit el-table--border el-table--enable-row-hover el-table--enable-row-transition"]
        can = wait_return_subelement_absolute(wait, 0.5, "//main[@class ='el-main main']")
        can_html = can.get_attribute('outerHTML')
        soup = BeautifulSoup(can_html, 'html.parser')
        result = soup.find('span', class_ = "el-table__empty-text")
        if result:
            return 0
        else:
            return -1










def schedule(complete, total, xpath, wait, member_excel, member_excel_path, control, wb, ws):
    global driver0
    global amount_mem_complete
    global amount_infomem_complete
    global amount_activist_complete
    global amount_devtar_complete
    global amount_applicant_complete
    global temp_countx
    #根据控制信息切换人员类型
    if control == 1:
        switch_formal_mem(wait)
    elif control == 2:
        switch_informal_mem(wait)
    elif control == 3:
        switch_devtarg(wait)
    elif control == 4:
        switch_activist(wait)
    elif control == 5:
        switch_applicant(wait)
    #设置每一张页面显示个人信息的条数
    set_amount_perpage(wait)
    #初始化初始页面为1
    page_number_old = 1
    #检测总完成条目是否小于total
    while complete+1 <= total:
        #检查即将录取的条目是否是100的整数
        row_number = int((complete+1) % 100)
        #计算得到页码和行码
        if row_number == 0:
            page_number = int((complete + 1 )/ 100)
            row_number = 100
        else:
            page_number = int((complete + 1)/ 100) + 1
            row_number = int((complete + 1) % 100)
        #time.sleep(0.1) 
        #判断是否需要改变页码
        if page_number != page_number_old:
            input_page = wait.until(EC.visibility_of_element_located((By.XPATH, xpath)))
            input_page.send_keys(Keys.CONTROL + "a")
            input_page.send_keys(Keys.BACKSPACE)
            input_page.send_keys(page_number)
            input_page.send_keys(Keys.RETURN)
            time.sleep(1)
        else:
            pass
        #进入个人页面函数
        yixianrudang = enter_person_infopage(wait, driver0, row_number, member_excel, member_excel_path, page_number, xpath="//li[@class = 'el-select-dropdown__item selected hover']//span", control = control, temp_countx = temp_countx, wb=wb, ws=ws)
        #逐条录入个人信息项到excel
        downloading(file = member_excel, wait = wait, path = member_excel_path, control = control, wb=wb, ws=ws, yixianrudang = yixianrudang)
        #录入完毕，重新回到选人页面
        driver0.close()
        try:
            for handle in driver0.window_handles:
                driver0.switch_to.window(handle)
                if '发展党员纪实' in driver0.title:
                    break
        except:
            time.sleep(0.3)
        #在这里检查线程关闭信号
        if stop_event.is_set():
            break
        #更新完成总数
        if control == 1:
            complete = amount_mem_complete
        elif control == 2:
            complete = amount_infomem_complete
        elif control == 3:
            complete = amount_devtar_complete
        elif control == 4:
            complete = amount_activist_complete
        elif control == 5:
            complete = amount_applicant_complete
        #更新选人页面老页码
        page_number_old = page_number



def access_info_page(wait, rowx, file, path):
    global temp_countx
    lishudangzuzhi = wait_return_subelement_absolute(wait, time_w = 0.5, xpath = f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[4]//a").text
    #lishudangzuzhi = wait.until(EC.visibility_of_element_located((By.XPATH, f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[5]//a"))).text
    file.active.cell(row=temp_countx+2, column=28).value = lishudangzuzhi
    file.save(path)
    wait_click_xpath(wait, time_w = 0.5, xpath = f"//table[@class = 'el-table__body']/tbody/tr[{rowx}]/td[1]//a")
    
    
    




def downloading(file, wait, path, control, wb, ws, yixianrudang = False):
    global driver0
    global amount_mem_complete
    global amount_infomem_complete
    global amount_activist_complete
    global amount_devtar_complete
    global amount_applicant_complete
    global temp_countx
    if control == 1:
        count = amount_mem_complete
    elif control == 2:
        count = amount_infomem_complete + amount_mem_complete
    elif control == 3:
        count = amount_devtar_complete + amount_infomem_complete + amount_mem_complete
    elif control == 4:
        count = amount_devtar_complete + amount_activist_complete + amount_infomem_complete + amount_mem_complete
    elif control == 5:
        count = amount_applicant_complete + amount_devtar_complete + amount_activist_complete + amount_infomem_complete + amount_mem_complete
        

    if control == 1:
        if yixianrudang == False:
            count = count + 1
            countx = count + 1
            temp_countx = count
            name_temp = jibenxinxi_download(file, path, wait, countx, count, wb, ws)
            jijifenzi_download(path, file, wait, countx, control, wb, ws)
            yubeidangyuan(path, file, wait, countx, control, wb, ws)
            ershiqi = "正式党员"
            ws.cell(row = countx, column = 27, value =ershiqi)
            print("填写第",count,"个正式党员",name_temp,"信息成功") 
            amount_mem_complete = amount_mem_complete + 1
        elif yixianrudang== True:
            count = count + 1
            countx = count + 1
            temp_countx = count
            name_temp = formal_yixian_download(file, path, wait, countx, count, wb, ws)
            ershiqi = "正式党员"
            ws.cell(row = countx, column = 27, value =ershiqi)
            print("填写第",count,"个正式党员",name_temp,"信息成功") 
            amount_mem_complete = amount_mem_complete + 1



    elif control == 2:
        count = count + 1
        countx = count + 1
        temp_countx = count
        name_temp = jibenxinxi_download(file, path, wait, countx, count, wb, ws)
        jijifenzi_download(path, file, wait, countx, control, wb, ws)
        yubeidangyuan(path, file, wait, countx, control, wb, ws)
        ershiqi = "预备党员"
        ws.cell(row = countx, column = 27, value =ershiqi)
        print("填写第",count,"个预备党员",name_temp,"信息成功") 
        amount_infomem_complete = amount_infomem_complete + 1



    elif control == 3:
        count = count + 1
        countx = count + 1
        temp_countx = count
        name_temp = jibenxinxi_download(file, path, wait, countx, count, wb, ws)
        jijifenzi_download(path, file, wait, countx, control, wb, ws)
        
        #加入党组织日期
        ershiyi = "-"
        ws.cell(row = countx, column = 21, value =ershiyi)
        #人员类别
        ershisan = "-"
        ws.cell(row = countx, column = 23, value =ershisan)
        #党籍状态
        ershisi = "-"
        ws.cell(row = countx, column = 24, value =ershisi)
        #入党类型
        ershiwu = "-"
        ws.cell(row = countx, column = 25, value =ershiwu)
        # 人员类型
        ershiqi = "发展对象"
        ws.cell(row = countx, column = 27, value =ershiqi)


        #加入党组织日期
        ershiyi = "-"
        ws.cell(row = countx, column = 21, value =ershiyi)
        #转为正式党员日期
        ershier = "-"
        ws.cell(row = countx, column = 22, value =ershier)
        #人员类别
        ershisan = "-"
        ws.cell(row = countx, column = 23, value =ershisan)
        #党籍状态
        ershisi = "-"
        ws.cell(row = countx, column = 24, value =ershisi)
        #入党类型
        ershiwu = "-"
        ws.cell(row = countx, column = 25, value =ershiwu)

        #预备党员阶段一线情况
        ershijiu = "-"
        ws.cell(row = countx, column = 29, value =ershijiu)

        #预备党员阶段工作岗位
        sanshi = "-"
        ws.cell(row = countx, column = 30, value =sanshi)

        #预备党员阶段所属行业
        sanshiyi = "-"
        ws.cell(row = countx, column = 31, value =sanshiyi)

        print("填写第",count,"个发展对象",name_temp,"信息成功") 
        amount_devtar_complete = amount_devtar_complete + 1



    elif control == 4:
        count = count + 1
        countx = count + 1
        temp_countx = count
        name_temp = jibenxinxi_download(file, path, wait, countx, count, wb, ws)
        jijifenzi_download(path, file, wait, countx, control, wb, ws)
        #加入党组织日期
        ershiyi = "-"
        ws.cell(row = countx, column = 21, value =ershiyi)
        #转为正式党员日期
        ershier = "-"
        ws.cell(row = countx, column = 22, value =ershier)
        #人员类别
        ershisan = "-"
        ws.cell(row = countx, column = 23, value =ershisan)
        #党籍状态
        ershisi = "-"
        ws.cell(row = countx, column = 24, value =ershisi)
        #入党类型
        ershiwu = "-"
        ws.cell(row = countx, column = 25, value =ershiwu)

        # 人员类型
        ershiqi = "积极分子"
        ws.cell(row = countx, column = 27, value =ershiqi)

        #预备党员阶段一线情况
        ershijiu = "-"
        ws.cell(row = countx, column = 29, value =ershijiu)

        #预备党员阶段工作岗位
        sanshi = "-"
        ws.cell(row = countx, column = 30, value =sanshi)

        #预备党员阶段所属行业
        sanshiyi = "-"
        ws.cell(row = countx, column = 31, value =sanshiyi)

        print("填写第",count,"个积极分子",name_temp,"信息成功") 
        amount_activist_complete = amount_activist_complete + 1



    elif control == 5:
        count = count + 1
        countx = count + 1
        temp_countx = count
        name_temp = jibenxinxi_download(file, path, wait, countx, count, wb, ws)
        if "入党积极分子基本信息" in driver0.page_source:
            jijifenzi_download(path, file, wait, countx, control, wb, ws)
            # 加入党组织日期
            ershiyi = "-"
            ws.cell(row = countx, column = 21, value =ershiyi)
            # 转为正式党员日期
            ershier = "-"
            ws.cell(row = countx, column = 22, value =ershier)
            # 人员类别
            ershisan = "-"
            ws.cell(row = countx, column = 23, value =ershisan)
            # 党籍状态
            ershisi = "-"
            ws.cell(row = countx, column = 24, value =ershisi)
            # 入党类型
            ershiwu = "-"
            ws.cell(row = countx, column = 25, value =ershiwu)
        else:
            #籍贯
            shisi = "-"
            ws.cell(row = countx, column = 14, value =shisi)
            #入团日期
            shiwu = "-"
            ws.cell(row = countx, column = 15, value =shiwu)
            #参加工作日期
            shiliu = "-"
            ws.cell(row = countx, column = 16, value =shiliu)
            #申请入党日期
            shiqi = wait.until(EC.presence_of_element_located((By.XPATH, "//tbody/tr[3]/td[2]"))).text
            ws.cell(row = countx, column = 17, value =shiqi)
            #确定入党积极分子日期
            shiba = "-"
            ws.cell(row = countx, column = 18, value =shiba)
            #工作单位及职务
            shijiu = "-"
            ws.cell(row = countx, column = 19, value =shijiu)
            #家庭住址
            ershi = "-"
            ws.cell(row = countx, column = 20, value =ershi)
            # 加入党组织日期
            ershiyi = "-"
            ws.cell(row = countx, column = 21, value =ershiyi)
            # 转为正式党员日期
            ershier = "-"
            ws.cell(row = countx, column = 22, value =ershier)
            # 人员类别
            ershisan = "-"
            ws.cell(row = countx, column = 23, value =ershisan)    
            # 党籍状态
            ershisi = "-"
            ws.cell(row = countx, column = 24, value =ershisi)
            # 入党类型
            ershiwu = "-"
            ws.cell(row = countx, column = 25, value =ershiwu)

            # 积极分子阶段工作岗位
            sanshisan = "-"
            ws.cell(row = countx, column = 33, value =sanshisan)

            # 积极分子阶段一线情况
            ershiliu = "-"
            ws.cell(row = countx, column = 26, value = ershiliu)


        # 人员类型
        ershiqi = "入党申请人"
        ws.cell(row = countx, column = 27, value =ershiqi)

        #预备党员阶段一线情况
        ershijiu = "-"
        ws.cell(row = countx, column = 29, value =ershijiu)

        #预备党员阶段工作岗位
        sanshi = "-"
        ws.cell(row = countx, column = 30, value =sanshi)

        #预备党员阶段所属行业
        sanshiyi = "-"
        ws.cell(row = countx, column = 31, value =sanshiyi)
        
        print("填写第",count,"个入党申请人",name_temp,"信息成功") 
        amount_applicant_complete = amount_applicant_complete + 1



def wait_click_xpath(wait, time_w, xpath):
    while(1):
        try:
            button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            break
        except:
            time.sleep(time_w)
    while(1):
        try:
            button.click()
            break
        except Exception as e:
            print(f"An exception occurred: {e}")
            button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            time.sleep(time_w)



def wait_click_xpath_action(driver, wait, time_w, xpath):
    while(1):
        try:
            button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            break
        except:
            time.sleep(time_w)
    while(1):
        try:
            ActionChains(driver).move_to_element(button).click().perform()
            break
        except Exception as e:
            print(f"An exception occurred: {e}")
            button = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            time.sleep(time_w)
