import openpyxl


#给定xls的路径，单元格，返回列的内容到列表
def xls2list(workbookpath, str, list):
    # 打开指定的 Excel 文件
    workbook = openpyxl.load_workbook(workbookpath)
    sheet = workbook.active  # 默认选择活动的工作表
    # 指定单元格
    #start_cell = 'A1'
    start_cell = str
    start_row = sheet[start_cell].row
    list = []
    letter = ''.join([char for char in start_cell if char.isalpha()])
    number = char_to_number(letter)
    #number = ''.join([char for char in start_cell if char.isdigit()])
    # 遍历指定单元格下方的所有行
    for row in sheet.iter_rows(min_row=start_row, min_col=int(number), max_col=int(number)):
        for cell in row:
            if cell.value is not None:  # 检查单元格是否有内容
                list.append(cell.value)
    return list


def char_to_number(char):
    return ord(char.lower()) - ord('a') + 1

def list2xls(workbookpath, str, list):
    workbook = openpyxl.load_workbook(workbookpath)
    ws = workbook.active
    start_cell = str
    start_row = ws[start_cell].row
    letter = ''.join([char for char in start_cell if char.isalpha()])
    number = char_to_number(letter)

    for i, value in enumerate(list):
        ws.cell(row=start_row + i, column=number, value=value)
    
    workbook.save(workbookpath)


    