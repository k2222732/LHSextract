from openpyxl import load_workbook


def init_complete_amount(excel_file_path):
    row_count = count_non_empty_rows(excel_file_path, sheet_name=0)
    amount_that_complete = row_count - 1
    return amount_that_complete


def count_non_empty_rows(excel_file_path, sheet_name=0):
    # 加载Excel工作簿
    workbook = load_workbook(filename=excel_file_path, data_only=True)
    # 获取工作表（可以通过名称或索引获取）
    if isinstance(sheet_name, int):
        sheet = workbook.worksheets[sheet_name]
    else:
        sheet = workbook[sheet_name]
    non_empty_row_count = 0
    
    # 逐行检查是否有数据
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            non_empty_row_count += 1
    return non_empty_row_count