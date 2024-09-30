from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
import time
import os
import re
import inspect
import configparser
import threading
from openpyxl import load_workbook
import openpyxl
from base.membase import *
from base.waitclick import *
stop_event = threading.Event()
captcha = ""
amount_that_complete = 0
mem_directory = ""



def stop_member_thread():
    stop_event.set()


def access_member_database(driver, wait):
    Databaseofparty = wait.until(EC.element_to_be_clickable((By.XPATH, '(//img[contains(@src, "党组织和党员信息库.png")])[2]')))
    Databaseofparty.click()
    try:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if '党组织和党员信息' in driver.title:
                break
        print(f"进入党组织和党员信息库成功")
    except:
        time.sleep(1)



def switch_role(wait):
    
    while 1:
        try:
            droplist = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.avatar-wrapper.fs-dropdown-selfdefine')))
            droplist.click()
            print(f"进入角色下拉列表成功")
            time.sleep(1)
            #role = wait.until(EC.element_to_be_clickable((By.XPATH, '//SPAN[contains(text(), "中国共产党山东汶上经济开发区工作委员会-具有审批预备党员权限的基层党委管理员")]')))
            config_file = 'config.ini'
            config = configparser.ConfigParser()
            config.read(config_file)
            role_name_mem = config.get('role_mem_name', 'name_role_mem', fallback='')
            role = wait.until(EC.element_to_be_clickable((By.XPATH, f"//SPAN[contains(text(), '{role_name_mem}')]")))
            role.click()
            if wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.avatar-wrapper.fs-dropdown-selfdefine'))).text == f"{role_name_mem}":
                print(f"切换角色成功")
                break
            else:
                pass
        except:
            time.sleep(2)


def new_excel(driver, wait, member_total_amount):
    global mem_directory
    config = configparser.ConfigParser()
    config.read('config.ini')
    mem_directory = config.get('Paths_mem_info', 'mem_info_path')
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党员信息库.xlsx"
    excel_file_path = os.path.join(mem_directory, excel_file_name)
    if os.path.isfile(excel_file_path):
        member_excel = load_workbook(excel_file_path)
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        rebuild(driver, excel_file_path, wait, member_total_amount, member_excel, excel_file_path, wb, ws)
    else:
        #在当前文件夹新建一个文件夹命名为"database"在里面新建一个文件夹名为"database_member"
        os.makedirs(mem_directory, exist_ok=True)
        #在database_member文件夹里新建一个excel文件命名为"今天的日期"&"党员信息库"
        #在excel里的第一行建立表头，分别是"姓名、性别、公民身份号码、民族、出生日期、学历、人员类别、学位、所在党支部、手机号码、入党日期
        #转正日期、党龄、党龄校正值、新社会阶层类型、工作岗位、从事专业技术职务、是否农民工、现居住地、户籍所在地、是否失联党员、是否流动党员、
        #党员注册时间、注册手机号、党员增加信息、党员减少信息、入党类型、转正情况、入党时所在支部、延长预备期时间
        columns = ["序号","姓名", "性别", "公民身份号码", "民族", "出生日期", "学历", "人员类别", "学位", 
            "所在党支部", "手机号码", "入党日期", "转正日期", "党龄", "党龄校正值", "新社会阶层类型", 
            "工作岗位", "从事专业技术职务", "是否农民工", "现居住地", "户籍所在地", "是否失联党员", 
            "是否流动党员", "入党类型", "转正情况", "入党时所在支部", "延长预备期时间"]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "sheet1"
        for i, value in enumerate(columns, start=1):
            ws.cell(row=1, column = i, value=value)
        wb.save(excel_file_path)

        member_excel = load_workbook(excel_file_path)
        print(f"文件 '{excel_file_path}' 已成功创建。")
        synchronizing(driver, wait, member_total_amount, member_excel, excel_file_path, wb, ws)
  

def get_amountof_member(wait):
    time.sleep(1)
    char_amountof_member = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'page']//span[@class = 'fs-pagination__total']"))).text
    int_amountof_members = re.findall(r'\d+', char_amountof_member)
    int_amountof_member = int(int_amountof_members[0]) if int_amountof_members else None
    return int(int_amountof_member)


def set_amount_perpage(wait):
    time.sleep(0.5)
    list = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class = 'page']//input[@type = 'text']")))
    list.click()
    time.sleep(0.5)
    option = wait.until(EC.element_to_be_clickable((By.XPATH, "(//ul[@class = 'fs-scrollbar__view fs-select-dropdown__list'])[4]/li[6]")))
    option.click()

    #//div[@class = 'page']//input[@type = 'text']


def access_info_page(wait, row):
    xpath = f"(//table[@class='fs-table__body'])[3]/tbody/tr[{row}]/td[3]//div[@role = 'button']"
    while 1:
        try:
            member = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
            break
        except:
            time.sleep(0.1)
    while 1:
        try:
            member.click()
            break
        except:
                while 1:
                    try:
                        member = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
                        break
                    except:
                        time.sleep(0.1)
    print(f"进入党员个人页面成功") 


def rebuild(driver, excel_file_path, wait, member_total_amount, member_excel, member_excel_path, wb, ws):
    #先改变全局变量
    #相比dev_func的rebuild，mem_func的rebuild不需要重写，也不用预先写
    global amount_that_complete
    amount_that_complete = init_complete_amount(excel_file_path)
    synchronizing(driver, wait, member_total_amount, member_excel, member_excel_path, wb, ws)
    


def synchronizing(driver, wait, member_total_amount, member_excel, member_excel_path, wb, ws):
    global amount_that_complete
    while amount_that_complete < member_total_amount:
        page_number = int(amount_that_complete / 100 + 1)
        row_number = int(amount_that_complete % 100 + 1)
        #time.sleep(0.1)
        input_page = wait.until(EC.visibility_of_element_located((By.XPATH, "(//div[@class = 'page']//input)[2]")))
        input_page.send_keys(Keys.CONTROL + "a")
        input_page.send_keys(Keys.BACKSPACE)
        input_page.send_keys(page_number)
        input_page.send_keys(Keys.RETURN)

        scroll = wait_return_subelement_absolute_v2(wait, 1, "//div[@class = 'fs-table__body-wrapper is-scrolling-left' or @class = 'fs-table__body-wrapper is-scrolling-right']")
        scroll_height = scroll.get_attribute('scrollHeight')
        if member_total_amount - amount_that_complete < 100:
            strip_num = member_total_amount % 100
        else:
            strip_num = 100
        current_scroll = ((row_number%100)/strip_num)*int(scroll_height)
        access_info_page(wait, row_number)
        while 1:
            try:
                downloading(amount_that_complete, member_excel, wait, member_excel_path, wb, ws)
                driver.execute_script("arguments[0].scrollTop = arguments[1]", scroll, current_scroll)
                break
            except:
                access_info_page(wait, row_number)
        amount_that_complete = amount_that_complete + 1
        ##在这里检查线程关闭信号
        if stop_event.is_set():
            break
    wb.save(member_excel_path)


def downloading(count, file, wait, path, wb, ws):
    while 1:
            rylb = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[7]"))).text
            if rylb == '正式党员':
                downloading_formal(count, file, wait, path, wb, ws)
                break
            elif rylb == '预备党员':
                downloading_informal(count, file, wait, path, wb, ws)
                break
            else:
                print('既不是正式党员也不是预备党员')
        


def downloading_informal(count, file, wait, path, wb, ws):
    count = count + 1
    countx = count + 1
    #填写序号
    yi = count
    ws.cell(row=countx, column = 1, value = yi)
    #file.save(path)
    #填写姓名
    er = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[1]"))).text
    ws.cell(row=countx, column=2, value=er)
    #file.save(path)
    #性别
    san = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[3]"))).text
    ws.cell(row=countx, column = 3, value = san)
    #file.save(path)
    #身份证
    si = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[2]"))).text
    ws.cell(row=countx, column = 4, value = si)
    #file.save(path)
    #民族
    wu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[4]"))).text
    ws.cell(row=countx, column = 5, value = wu)
    #file.save(path)
    #出生日期
    liu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[5]"))).text
    ws.cell(row=countx, column = 6, value = liu)
    #file.save(path)
    #学历
    qi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[6]"))).text
    ws.cell(row=countx, column = 7, value = qi)
    #file.save(path)
    #人员类别
    ba = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[7]"))).text
    ws.cell(row=countx, column = 8, value = ba)
    #file.save(path)
    #学位
    jiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[8]"))).text
    ws.cell(row=countx, column = 9, value = jiu)
    #file.save(path)
    #所在党支部
    shi = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'card-class']//div[@class = 'row-vals-shot']"))).text
    ws.cell(row=countx, column = 10, value = shi)
    #file.save(path)
    #手机号码
    shiyi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[9]"))).text
    ws.cell(row=countx, column = 11, value = shiyi)
    #file.save(path)
    #入党日期
    shier = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[10]"))).text
    ws.cell(row=countx, column = 12, value = shier)
    #file.save(path)
    #转正日期
    shisan = '-'
    ws.cell(row=countx, column = 13, value = shisan)
    #file.save(path)
    #党龄
    shisi = '-'
    ws.cell(row=countx, column = 14, value = shisi)
    #file.save(path)
    #党龄矫正值
    shiwu = '-'
    ws.cell(row=countx, column = 15, value = shiwu)
    #file.save(path)
    #新社会阶层类型
    shiliu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[2]"))).text
    ws.cell(row=countx, column = 16, value = shiliu)
    #file.save(path)
    #工作岗位
    temp_job = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '工作岗位')])[1]/following-sibling::div[@class = 'row-val-shot']"))).text
    while 1:
        if temp_job == None or '':
            temp_job = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '工作岗位')])[1]/following-sibling::div[@class = 'row-val-shot']"))).text
        else:
            break
    shiqi = temp_job
    ws.cell(row=countx, column = 17, value = shiqi)
    #file.save(path)
    #从事专业技术职务
    shiba = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '从事专业技术职务')])[1]/following-sibling::div[@class = 'row-val-shot']"))).text
    ws.cell(row=countx, column = 18, value = shiba)
    #file.save(path)
    #是否农民工
    shijiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '是否农民工')])[1]/following-sibling::div[@class = 'row-val-big']"))).text
    ws.cell(row=countx, column = 19, value = shijiu)
    #file.save(path)
    #现居住地
    ershi = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '现居住地')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 20, value = ershi)
    #file.save(path)
    #户籍所在地
    ershiyi = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '户籍所在地')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 21, value = ershiyi)
    #file.save(path)
    #是否失联党员
    ershier = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '是否失联党员')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 22, value = ershier)
    #file.save(path)
    #是否流动党员
    ershisan = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '是否流动党员')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 23, value = ershisan)
    
    #切换选项卡
    switch_card_joininfo = wait.until(EC.element_to_be_clickable((By.ID, "tab-enterInfo")))
    switch_card_joininfo.click()
    #入党类型
    wait.until(EC.text_to_be_present_in_element((By.XPATH, "(//div[@class = 'table-row'])[1]/div[@class = 'row-key'][1]"), '入党类型'))
    ershisi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '入党类型')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 24, value = ershisi)
    #file.save(path)
    #转正情况
    ershiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '转正情况')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 25, value = ershiwu)
    #file.save(path)
    #入党时所在党支部
    ershiliu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '入党时所在支部')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 26, value = ershiliu)
    #file.save(path)
    #延长预备期时间
    ershiqi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '延长预备期时间')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 27, value = ershiqi)
    if count%100 == 0:
        wb.save(path)
    #debugging()
    print("填写第",count,"名预备党员",er,"信息成功") 
    exit_member_card = wait.until(EC.element_to_be_clickable((By.XPATH, "(//button[@class = 'fs-button fs-button--default fs-button--small'])[2]")))
    exit_member_card.click()
    


def downloading_formal(count, file, wait, path, wb, ws):
    count = count + 1
    countx = count + 1
    #填写序号
    yi = count
    ws.cell(row=countx, column = 1, value = yi)
    #file.save(path)
    #填写姓名
    er = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[1]"))).text
    ws.cell(row=countx, column = 2, value = er)
    #file.save(path)
    #性别
    san = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[3]"))).text
    ws.cell(row=countx, column = 3, value = san)
    #file.save(path)
    #身份证
    si = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[2]"))).text
    ws.cell(row=countx, column = 4, value = si)
    #file.save(path)
    #民族
    wu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[4]"))).text
    ws.cell(row=countx, column = 5, value = wu)
    #file.save(path)
    #出生日期
    liu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[5]"))).text
    ws.cell(row=countx, column = 6, value = liu)
    #file.save(path)
    #学历
    qi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[6]"))).text
    ws.cell(row=countx, column = 7, value = qi)
    #file.save(path)
    #人员类别
    ba = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[7]"))).text
    ws.cell(row=countx, column = 8, value = ba)
    #file.save(path)
    #学位
    jiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[8]"))).text
    ws.cell(row=countx, column = 9, value = jiu)
    #file.save(path)
    #所在党支部
    shi = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'card-class']//div[@class = 'row-vals-shot']"))).text
    ws.cell(row=countx, column = 10, value = shi)
    #file.save(path)
    #手机号码
    shiyi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[9]"))).text
    ws.cell(row=countx, column = 11, value = shiyi)
    #file.save(path)
    #入党日期
    shier = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[10]"))).text
    ws.cell(row=countx, column = 12, value = shier)
    #file.save(path)
    #转正日期
    shisan = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[1]"))).text
    ws.cell(row=countx, column = 13, value = shisan)
    #file.save(path)
    #党龄
    shisi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'table-row'])[7]//div[2]"))).text
    ws.cell(row=countx, column = 14, value = shisi)
    #file.save(path)
    #党龄矫正值
    shiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[2]"))).text
    ws.cell(row=countx, column = 15, value = shiwu)
    #file.save(path)
    #新社会阶层类型
    shiliu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[4]"))).text
    ws.cell(row=countx, column = 16, value = shiliu)
    #file.save(path)
    #工作岗位
    temp_job = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '工作岗位')])[1]/following-sibling::div[@class = 'row-val-shot']"))).text
    while 1:
        if temp_job == None or '':
            temp_job = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '工作岗位')])[1]/following-sibling::div[@class = 'row-val-shot']"))).text
        else:
            break
    shiqi = temp_job
    ws.cell(row=countx, column = 17, value = shiqi)
    #file.save(path)
    #从事专业技术职务
    shiba = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '从事专业技术职务')])[1]/following-sibling::div[@class = 'row-val-shot']"))).text
    ws.cell(row=countx, column = 18, value = shiba)
    #file.save(path)
    #是否农民工
    shijiu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '是否农民工')])[1]/following-sibling::div"))).text
    ws.cell(row=countx, column = 19, value = shijiu)
    #file.save(path)
    #现居住地
    ershi = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '现居住地')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 20, value = ershi)
    #file.save(path)
    #户籍所在地
    ershiyi = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '户籍所在地')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 21, value = ershiyi)
    #file.save(path)
    #是否失联党员
    ershier = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '是否失联党员')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 22, value = ershier)
    #file.save(path)
    #是否流动党员
    ershisan = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[contains(text(), '是否流动党员')])[1]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 23, value = ershisan)
    
    #切换选项卡
    switch_card_joininfo = wait.until(EC.element_to_be_clickable((By.ID, "tab-enterInfo")))
    switch_card_joininfo.click()
    #入党类型
    wait.until(EC.text_to_be_present_in_element((By.XPATH, "(//div[@class = 'table-row'])[1]/div[@class = 'row-key'][1]"), '入党类型'))
    ershisi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '入党类型')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 24, value = ershisi)
    #file.save(path)
    #转正情况
    ershiwu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '转正情况')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 25, value = ershiwu)
    #file.save(path)
    #入党时所在党支部
    ershiliu = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '入党时所在支部')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 26, value = ershiliu)
    #file.save(path)
    #延长预备期时间
    ershiqi = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[contains(text(), '延长预备期时间')]/following-sibling::div)[1]"))).text
    ws.cell(row=countx, column = 27, value = ershiqi)
    if count%100 == 0:
        wb.save(path)
    #debugging()
    print("填写第",count,"名党员",er,"信息成功") 
    exit_member_card = wait.until(EC.element_to_be_clickable((By.XPATH, "(//button[@class = 'fs-button fs-button--default fs-button--small'])[2]")))
    exit_member_card.click()


