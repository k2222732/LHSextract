from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from selenium import webdriver
from datetime import datetime
import pandas as pd
import time 
import os
import re
import inspect

amount_that_complete = 0




def driver_create(chrome_path, chromedriver_path):
    chrome_options = Options()
    # 替换为你的 Chrome 浏览器的实际安装路径
    chrome_options.binary_location = chrome_path  
    # 创建 Service 对象并指定 ChromeDriver 的路径
    service = Service(executable_path=chromedriver_path)
    # 启动 WebDriver
    driver = webdriver.Chrome(service=service, options=chrome_options)
    #返回driver
    print(f"浏览器驱动创建成功")
    return driver



def login(account, password, driver, url, wait):
    # 打开网址
    driver.get(url)
    username_box = wait.until(EC.visibility_of_element_located((By.ID, 'username')))
    password_box = wait.until(EC.visibility_of_element_located((By.ID, 'password')))
    validatecode = wait.until(EC.visibility_of_element_located((By.ID, 'validateCode')))
    login_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.js-submit.tianze-loginbtn')))
    username_box.send_keys(account)
    password_box.send_keys(password)
    # 等待手动输入验证码
    temp = input ("Please enter the captcha and hit enter in the browser")
    validatecode.send_keys(temp)
    # 点击登录按钮
    login_button.click()
    # 之后可以添加额外的代码来处理登录后的页面或关闭浏览器
    print(f"登录成功")


def access_member_database(driver, wait):
    Databaseofparty = wait.until(EC.element_to_be_clickable((By.XPATH, '(//img[contains(@src, "党组织和党员信息库.png")])[2]')))
    Databaseofparty.click()
    try:
        for handle in driver.window_handles:
            driver.switch_to.window(handle)
            if '党组织和党员信息' in driver.title:
                break
        print(f"进入党组织和党员信息库成功")
    except:
        time.sleep(1)



def switch_role(wait):
    droplist = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '.avatar-wrapper.fs-dropdown-selfdefine')))
    droplist.click()
    print(f"进入角色下拉列表成功")
    time.sleep(3)
    role = wait.until(EC.element_to_be_clickable((By.XPATH, '//SPAN[contains(text(), "中国共产党山东汶上经济开发区工作委员会-具有审批预备党员权限的基层党委管理员")]')))
    role.click()
    print(f"切换角色成功")


def new_excel():
    #在当前文件夹新建一个文件夹命名为"database"在里面新建一个文件夹名为"database_member"
    os.makedirs("database/database_member", exist_ok=True)
    #在database_member文件夹里新建一个excel文件命名为"今天的日期"&"党员信息库"
    today_date = datetime.now().strftime("%Y-%m-%d")
    excel_file_name = f"{today_date}党员信息库.xlsx"
    excel_file_path = os.path.join("database", "database_member", excel_file_name)
    #在excel里的第一行建立表头，分别是"姓名、性别、公民身份号码、民族、出生日期、学历、人员类别、学位、所在党支部、手机号码、入党日期
    #转正日期、党龄、党龄校正值、新社会阶层类型、工作岗位、从事专业技术职务、是否农民工、现居住地、户籍所在地、是否失联党员、是否流动党员、
    #党员注册时间、注册手机号、党员增加信息、党员减少信息、入党类型、转正情况、入党时所在支部、延长预备期时间
    columns = ["序号","姓名", "性别", "公民身份号码", "民族", "出生日期", "学历", "人员类别", "学位", 
           "所在党支部", "手机号码", "入党日期", "转正日期", "党龄", "党龄校正值", "新社会阶层类型", 
           "工作岗位", "从事专业技术职务", "是否农民工", "现居住地", "户籍所在地", "是否失联党员", 
           "是否流动党员", "入党类型", "转正情况", "入党时所在支部", "延长预备期时间"]
    
    df = pd.DataFrame(columns=columns)
    df.to_excel(excel_file_path, index=False)

    workbook = load_workbook(excel_file_path)
    print(f"文件 '{excel_file_path}' 已成功创建。")
    return workbook, excel_file_path
  


def get_amountof_member(wait):
    #//div[@class = 'page']//span[@class = 'fs-pagination__total'] 
    char_amountof_member = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'page']//span[@class = 'fs-pagination__total']"))).text
    int_amountof_members = re.findall(r'\d+', char_amountof_member)
    int_amountof_member = int(int_amountof_members[0]) if int_amountof_members else None
    return int(int_amountof_member)


def set_amount_perpage(wait):
    list = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class = 'page']//input[@type = 'text']")))
    list.click()
    option = wait.until(EC.element_to_be_clickable((By.XPATH, "//li[@class = 'fs-select-dropdown__item hover']")))
    option.click()

    #//div[@class = 'page']//input[@type = 'text']

def access_info_page(wait, row):
    xpath = f"(//table[@class='fs-table__body'])[3]/tbody/tr[{row}]/td[3]"
    wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
    member = wait.until(EC.element_to_be_clickable((By.XPATH, xpath)))
    member.click() 
    print(f"进入党员个人页面成功") 

def synchronizing(wait, member_total_amount, member_excel, member_excel_path):
    global amount_that_complete
    while amount_that_complete < member_total_amount:
        page_number = amount_that_complete / 100 + 1
        row_number = amount_that_complete % 100 + 2
        #(//div[@class = 'page']//input)[2]
        input_page = wait.until(EC.visibility_of_element_located((By.XPATH, "(//div[@class = 'page']//input)[2]")))
        input_page.send_keys(page_number)
        access_info_page(wait, row_number)
        downloading(amount_that_complete, member_excel, wait, member_excel_path)
        amount_that_complete = amount_that_complete + 1




def downloading(count, file, wait, path):

    rylb = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[7]"))).text
    if rylb == '正式党员':
        downloading_formal(count, file, wait, path)
    elif rylb == '预备党员':
        downloading_informal(count, file, wait, path)
    else:
        print('既不是正式党员也不是预备党员')

        



def downloading_informal(count, file, wait, path):
    countx = count + 1
    #填写序号
    file.active.cell(row=countx, column=1).value = count
    file.save(path)
    #填写姓名
    name_temp = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[1]"))).text
    file.active.cell(row=countx, column=2).value = name_temp
    file.save(path)
    #性别
    file.active.cell(row=countx, column=3).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[3]"))).text
    file.save(path)
    #身份证
    file.active.cell(row=countx, column=4).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[2]"))).text
    file.save(path)
    #民族
    file.active.cell(row=countx, column=5).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[4]"))).text
    file.save(path)
    #出生日期
    file.active.cell(row=countx, column=6).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[5]"))).text
    file.save(path)
    #学历
    file.active.cell(row=countx, column=7).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[6]"))).text
    file.save(path)
    #人员类别
    file.active.cell(row=countx, column=8).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[7]"))).text
    file.save(path)
    #学位
    file.active.cell(row=countx, column=9).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[8]"))).text
    file.save(path)
    #所在党支部
    file.active.cell(row=countx, column=10).value = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'card-class']//div[@class = 'row-vals-shot']"))).text
    file.save(path)
    #手机号码
    file.active.cell(row=countx, column=11).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[9]"))).text
    file.save(path)
    #入党日期
    file.active.cell(row=countx, column=12).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[10]"))).text
    file.save(path)

    #转正日期
    file.active.cell(row=countx, column=13).value = '-'
    file.save(path)
    #党龄
    file.active.cell(row=countx, column=14).value = '-'
    file.save(path)
    #党龄矫正值
    file.active.cell(row=countx, column=15).value = '-'
    file.save(path)

    #新社会阶层类型
    file.active.cell(row=countx, column=17).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[2]"))).text
    file.save(path)
    #工作岗位
    file.active.cell(row=countx, column=16).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[1]"))).text
    file.save(path)
    #从事专业技术职务
    file.active.cell(row=countx, column=19).value = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[@class = 'table-row'])[8]//div[@class= 'select-dict'])[1]"))).text
    file.save(path)
    #是否农民工
    file.active.cell(row=countx, column=18).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'table-row'])[9]//span[@style = 'margin-top: auto; margin-bottom: auto;']"))).text
    file.save(path)
    #现居住地
    file.active.cell(row=countx, column=20).value = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[@class = 'card-class']//div[@class = 'table-row'])/div[@class = 'row-vals'])[1]"))).text
    file.save(path)
    #户籍所在地
    file.active.cell(row=countx, column=21).value = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[@class = 'card-class']//div[@class = 'table-row'])/div[@class = 'row-vals'])[2]"))).text
    file.save(path)
    #是否失联党员
    file.active.cell(row=countx, column=22).value = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[@class = 'card-class']//div[@class = 'table-row'])/div[@class = 'row-vals'])[3]"))).text
    file.save(path)
    #是否流动党员
    file.active.cell(row=countx, column=23).value = wait.until(EC.presence_of_element_located((By.XPATH, "((//div[@class = 'card-class']//div[@class = 'table-row'])/div[@class = 'row-vals'])[4]"))).text
    file.save(path)
    #切换选项卡
    switch_card_joininfo = wait.until(EC.element_to_be_clickable((By.ID, "tab-enterInfo")))
    switch_card_joininfo.click()
    #入党类型
    wait.until(EC.text_to_be_present_in_element((By.XPATH, "(//div[@class = 'table-row'])[1]/div[@class = 'row-key'][1]"), '入党类型'))
    file.active.cell(row=countx, column=24).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[1]"))).text
    file.save(path)
    #转正情况
    file.active.cell(row=countx, column=25).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[3]"))).text
    file.save(path)
    #入党时所在党支部
    file.active.cell(row=countx, column=26).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[5]"))).text
    file.save(path)
    #延长预备期时间
    file.active.cell(row=countx, column=27).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[6]"))).text
    file.save(path)
    #debugging()
    print("填写第",count,"名预备党员",name_temp,"信息成功") 
    exit_member_card = wait.until(EC.element_to_be_clickable((By.XPATH, "(//button[@class = 'fs-button fs-button--default fs-button--small'])[2]")))
    exit_member_card.click()



def downloading_formal(count, file, wait, path):
    countx = count + 1
    #填写序号
    file.active.cell(row=countx, column=1).value = count
    file.save(path)
    #填写姓名
    name_temp = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[1]"))).text
    file.active.cell(row=countx, column=2).value = name_temp
    file.save(path)
    #性别
    file.active.cell(row=countx, column=3).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[3]"))).text
    file.save(path)
    #身份证
    file.active.cell(row=countx, column=4).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[2]"))).text
    file.save(path)
    #民族
    file.active.cell(row=countx, column=5).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[4]"))).text
    file.save(path)
    #出生日期
    file.active.cell(row=countx, column=6).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[5]"))).text
    file.save(path)
    #学历
    file.active.cell(row=countx, column=7).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[6]"))).text
    file.save(path)
    #人员类别
    file.active.cell(row=countx, column=8).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[7]"))).text
    file.save(path)
    #学位
    file.active.cell(row=countx, column=9).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[8]"))).text
    file.save(path)
    #所在党支部
    file.active.cell(row=countx, column=10).value = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@class = 'card-class']//div[@class = 'row-vals-shot']"))).text
    file.save(path)
    #手机号码
    file.active.cell(row=countx, column=11).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[9]"))).text
    file.save(path)
    #入党日期
    file.active.cell(row=countx, column=12).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[10]"))).text
    file.save(path)
    #转正日期
    file.active.cell(row=countx, column=13).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[1]"))).text
    file.save(path)
    #党龄
    file.active.cell(row=countx, column=14).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'table-row'])[7]//div[2]"))).text
    file.save(path)
    #党龄矫正值
    file.active.cell(row=countx, column=15).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[2]"))).text
    file.save(path)
    #新社会阶层类型
    file.active.cell(row=countx, column=17).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[3]"))).text
    file.save(path)
    #工作岗位
    file.active.cell(row=countx, column=16).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[4]"))).text
    file.save(path)
    #从事专业技术职务
    file.active.cell(row=countx, column=19).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-big'])[5]"))).text
    file.save(path)
    #是否农民工
    file.active.cell(row=countx, column=18).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'fs-tabs__content']//div[@class = 'row-val-shot'])[11]"))).text
    file.save(path)
    #现居住地
    file.active.cell(row=countx, column=20).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'table-row'])[11]/div[@class = 'row-vals']"))).text
    file.save(path)
    #户籍所在地
    file.active.cell(row=countx, column=21).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'table-row'])[12]/div[@class = 'row-vals']"))).text
    file.save(path)
    #是否失联党员
    file.active.cell(row=countx, column=22).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'table-row'])[13]/div[@class = 'row-vals']"))).text
    file.save(path)
    #是否流动党员
    file.active.cell(row=countx, column=23).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'card-class']//div[@class = 'table-row'])[14]/div[@class = 'row-vals']"))).text
    file.save(path)
    #切换选项卡
    switch_card_joininfo = wait.until(EC.element_to_be_clickable((By.ID, "tab-enterInfo")))
    switch_card_joininfo.click()
    #入党类型
    wait.until(EC.text_to_be_present_in_element((By.XPATH, "(//div[@class = 'table-row'])[1]/div[@class = 'row-key'][1]"), '入党类型'))
    file.active.cell(row=countx, column=24).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[1]"))).text
    file.save(path)
    #转正情况
    file.active.cell(row=countx, column=25).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[3]"))).text
    file.save(path)
    #入党时所在党支部
    file.active.cell(row=countx, column=26).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[5]"))).text
    file.save(path)
    #延长预备期时间
    file.active.cell(row=countx, column=27).value = wait.until(EC.presence_of_element_located((By.XPATH, "(//div[@class = 'row-val'])[6]"))).text
    file.save(path)
    #debugging()
    print("填写第",count,"名党员",name_temp,"信息成功") 
    exit_member_card = wait.until(EC.element_to_be_clickable((By.XPATH, "(//button[@class = 'fs-button fs-button--default fs-button--small'])[2]")))
    exit_member_card.click()




def current_line_number():
    return inspect.currentframe().f_back.f_lineno


def debugging():
    current_file_path = __file__
    current_file_name = os.path.basename(current_file_path)
    print("Current line number:", current_file_name, current_line_number())


